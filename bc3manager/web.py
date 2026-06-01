"""
BC3Manager — Interfaz web local con edición inline.

Uso:  python -m bc3manager.web
"""
from __future__ import annotations
import os, sys, tempfile, webbrowser, json as _json
from threading import Timer
from typing import Optional
from flask import Flask, render_template_string, request, jsonify, send_file, Response
from bc3manager.core.model import Presupuesto, TipoConcepto, Hijo, Concepto
from bc3manager.io.lector import leer_bc3
from bc3manager.io.escritor import escribir_bc3
from bc3manager.reports.informes import generar_informe, INFORMES

app = Flask(__name__)
_estado: dict = {"presupuesto": None, "ruta_original": None}

def _fmt(v: float, dec: int = 2) -> str:
    s = f"{v:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return s

def _arbol_json(p: Presupuesto) -> list[dict]:
    if not p.codigo_raiz: return []
    def _nodo(codigo, codigo_padre):
        c = p.get(codigo)
        if c is None: return None
        hijos_data = []
        if c.tipo == TipoConcepto.CAPITULO:
            for h in c.hijos:
                hc = p.get(h.codigo_hijo)
                if hc and hc.tipo in (TipoConcepto.CAPITULO, TipoConcepto.PARTIDA):
                    n = _nodo(h.codigo_hijo, codigo)
                    if n: hijos_data.append(n)
        med_total = p.medicion_total(codigo, codigo_padre)
        importe = p._importe_recursivo(codigo, codigo_padre) if codigo_padre else p.presupuesto_total()
        recursos = []
        if c.tipo == TipoConcepto.PARTIDA:
            acum = 0.0   # acumulado para resolver líneas de porcentaje
            for h in c.hijos:
                rec = p.get(h.codigo_hijo)
                if rec:
                    es_pct = Presupuesto.es_porcentaje(rec)
                    if es_pct:
                        # Importe = (acumulado de líneas previas) × coeficiente.
                        # El "precio" mostrado es el % (p.ej. 7), no el precio unitario.
                        imp = round(acum * h.cantidad, 4)
                    else:
                        imp = round(rec.precio * h.cantidad, 4)
                    acum += imp
                    recursos.append({"codigo": rec.codigo, "unidad": rec.unidad, "resumen": rec.resumen,
                        "tipo_fiebdc": getattr(rec, "_tipo_fiebdc", "3"),
                        "es_porcentaje": es_pct,
                        "precio": rec.precio, "precio_fmt": _fmt(rec.precio, 4),
                        "rendimiento": h.rendimiento, "rendimiento_fmt": _fmt(h.rendimiento, 4),
                        "importe": imp,
                        "importe_fmt": _fmt(imp, 4)})
        lineas_med = []
        med_obj = c.mediciones.get(codigo_padre)
        if med_obj:
            for ln in med_obj.lineas:
                lineas_med.append({"comentario": ln.comentario, "n_uds": ln.n_uds,
                    "longitud": ln.longitud, "anchura": ln.anchura, "altura": ln.altura,
                    "subtotal": round(ln.subtotal, 3), "subtotal_fmt": _fmt(ln.subtotal, 3)})
        return {"codigo": c.codigo, "unidad": c.unidad, "resumen": c.resumen, "texto": c.texto,
            "tipo": c.tipo.value, "tipo_fiebdc": getattr(c, "_tipo_fiebdc", ""),
            "precio": c.precio, "precio_fmt": _fmt(c.precio, 2),
            "medicion": med_total, "medicion_fmt": _fmt(med_total, 2),
            "importe": importe, "importe_fmt": _fmt(importe, 2),
            "hijos": hijos_data, "recursos": recursos, "lineas_medicion": lineas_med,
            "padre": codigo_padre}
    raiz = p.get(p.codigo_raiz)
    if not raiz: return []
    return [n for h in raiz.hijos if (n := _nodo(h.codigo_hijo, p.codigo_raiz))]

def _info_json(p):
    raiz = p.get(p.codigo_raiz) if p.codigo_raiz else None
    # Detectar si el archivo está en carpeta temporal (subido por web, no abierto desde disco)
    ruta = _estado.get("ruta_original") or ""
    es_temporal = False
    if ruta:
        try:
            es_temporal = os.path.commonpath([ruta, tempfile.gettempdir()]) == tempfile.gettempdir()
        except (ValueError, OSError):
            es_temporal = False
    return {"obra": raiz.resumen if raiz else "Sin nombre", "version": p.version_formato,
        "programa": p.programa_emisor,
        "capitulos": sum(1 for c in p.conceptos.values() if c.tipo == TipoConcepto.CAPITULO),
        "partidas": sum(1 for c in p.conceptos.values() if c.tipo == TipoConcepto.PARTIDA),
        "unitarios": sum(1 for c in p.conceptos.values() if c.tipo == TipoConcepto.UNITARIO),
        "total": p.presupuesto_total(), "total_fmt": _fmt(p.presupuesto_total()),
        "archivo_temporal": es_temporal}

def _resp():
    p = _estado["presupuesto"]
    p.recalcular()          # garantiza precios actualizados tras cualquier edición
    pila = _estado.get("undo_stack")
    redo = _estado.get("redo_stack")
    undo_disponible = bool(pila and len(pila) >= 2)
    redo_disponible = bool(redo)
    return jsonify({"ok": True, "info": _info_json(p), "arbol": _arbol_json(p),
                    "undo_disponible": undo_disponible,
                    "redo_disponible": redo_disponible})


# ============ API ENDPOINTS ============
@app.route("/")
def index(): return render_template_string(HTML_TEMPLATE)

@app.route("/api/cargar", methods=["POST"])
def api_cargar():
    archivo = request.files.get("archivo")
    if not archivo: return jsonify({"error": "No se recibió archivo"}), 400
    # Guardar en temporal para la carga inicial; la ruta real vendrá del nombre original
    tmp = tempfile.NamedTemporaryFile(suffix=".bc3", delete=False)
    archivo.save(tmp.name); tmp.close()
    try:
        p = leer_bc3(tmp.name)
        _estado["presupuesto"] = p
        _estado["ruta_original"] = tmp.name
        _estado["nombre_archivo"] = archivo.filename or "presupuesto.bc3"
        return jsonify(_payload_carga(p))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cargar_local", methods=["POST"])
def api_cargar_local():
    """Carga un archivo por ruta local (para abrir desde argumento de línea de comandos)."""
    ruta = _estado.get("ruta_arg")
    if not ruta: return jsonify({"error": "No hay archivo para cargar"}), 400
    try:
        p = leer_bc3(ruta)
        _estado["presupuesto"] = p
        _estado["ruta_original"] = ruta
        _estado["nombre_archivo"] = os.path.basename(ruta)
        return jsonify(_payload_carga(p))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _payload_carga(p) -> dict:
    """Construye la respuesta JSON común a las dos rutas de carga:
    árbol + info + validación (precios, PEM) + Excel de validación.
    Hace UNA sola validación completa y la reutiliza."""
    val = p.validar_completo()                 # recalcula y compara todo
    _estado["discrepancias"] = val["precios"]
    _init_historial()                          # punto de partida para deshacer
    nombre = _estado.get("nombre_archivo", "presupuesto.bc3")
    # Volcado a consola
    _log_discrepancias(p, nombre)
    # Excel automático
    ruta_xlsx = _generar_validacion_xlsx_seguro(p)
    _estado["ruta_validacion_xlsx"] = ruta_xlsx
    if ruta_xlsx:
        _safe_print(f"[OK] Excel de validacion generado en: {ruta_xlsx}")
        _safe_print()
    # PEM siempre presente (números reales aunque coincidan)
    comp = p.comparar_importes_archivo()
    pem = {
        "archivo": comp["pem_archivo"],
        "calculado": comp["pem_calculado"],
        "diferencia": comp["diferencia_pem"],
    }
    return {"info": _info_json(p), "arbol": _arbol_json(p),
            "archivo": nombre,
            "discrepancias": val["precios"],
            "mediciones_inconsistentes": len(val["mediciones"]),
            "pem": pem,
            "validacion_xlsx": ruta_xlsx}

@app.route("/api/tiene_archivo")
def api_tiene_archivo():
    """Devuelve si hay un archivo pasado por argumento para carga automática."""
    return jsonify({"tiene": bool(_estado.get("ruta_arg"))})

def _safe_print(*args, **kwargs) -> None:
    """print() tolerante con consolas que no soportan UTF-8 (Windows cp1252)."""
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        # Sustituye caracteres no representables por '?' antes de imprimir
        enc = getattr(sys.stdout, "encoding", "utf-8") or "utf-8"
        line = " ".join(str(a) for a in args).encode(enc, errors="replace").decode(enc)
        print(line, **kwargs)


def _log_discrepancias(p, nombre_archivo: str) -> None:
    """Vuelca a consola TODAS las discrepancias entre el archivo BC3 y la
    propagación recalculada (mediciones, precios, PEM).
    Se ejecuta UNA vez al abrir el presupuesto."""
    val = p.validar_completo()
    sep = "=" * 78
    subsep = "-" * 78
    resumen = val["resumen"]

    _safe_print()
    _safe_print(sep)
    _safe_print(f"  VALIDACION COMPLETA - {nombre_archivo}")
    _safe_print(f"  Conceptos: {resumen['total_conceptos']}  "
                f"Mediciones: {resumen['total_mediciones']}  "
                f"PEM (propagado): {_fmt(resumen['pem_calculado'])} EUR")
    _safe_print(sep)

    n_med = len(val["mediciones"])
    n_pre = len(val["precios"])
    n_pem = 1 if val["pem"] else 0
    total = n_med + n_pre + n_pem

    if total == 0:
        _safe_print("  [OK] La propagacion coincide con lo declarado en el archivo:")
        _safe_print("       - Sumas de lineas de medicion coinciden con totales declarados")
        _safe_print("       - Precios de partidas/capitulos coinciden con descomposicion")
        _safe_print("       - PEM agregado coincide con suma de capitulos raiz")
        _safe_print(sep)
        _safe_print()
        return

    # ---- 1) Mediciones ----------------------------------------------------
    _safe_print()
    _safe_print(f"  [1] MEDICIONES  ({n_med} inconsistencia(s))")
    _safe_print(subsep)
    if not val["mediciones"]:
        _safe_print("     OK. Suma de lineas == total declarado en todos los registros ~M.")
    else:
        _safe_print(f"     {'PARTIDA':<14} {'EN':<14} {'DECLARADA':>14}  "
                    f"{'SUMA LINEAS':>14}  {'DIF':>10}")
        _safe_print(f"     {'-'*14} {'-'*14} {'-'*14}  {'-'*14}  {'-'*10}")
        for d in val["mediciones"][:30]:
            _safe_print(
                f"     {d['codigo']:<14} {d['padre']:<14} "
                f"{_fmt(d['declarada'], 4):>14}  "
                f"{_fmt(d['suma_lineas'], 4):>14}  "
                f"{_fmt(d['diferencia'], 4):>10}"
            )
        if len(val["mediciones"]) > 30:
            _safe_print(f"     ... y {len(val['mediciones']) - 30} mas")

    # ---- 2) Precios -------------------------------------------------------
    _safe_print()
    _safe_print(f"  [2] PRECIOS  ({n_pre} inconsistencia(s))")
    _safe_print(subsep)
    if not val["precios"]:
        _safe_print("     OK. Precios del archivo == precios recalculados desde descomposicion.")
    else:
        _safe_print(f"     {'CODIGO':<14} {'TIPO':<5} {'ARCHIVO':>14}  "
                    f"{'CALCULADO':>14}  {'DIF':>10}  {'%':>7}  DESCRIPCION")
        _safe_print(f"     {'-'*14} {'-'*5} {'-'*14}  {'-'*14}  {'-'*10}  {'-'*7}  {'-'*30}")
        for d in val["precios"][:30]:
            tipo = "CAP" if d["tipo"] == "capitulo" else "PAR"
            _safe_print(
                f"     {d['codigo']:<14} {tipo:<5} "
                f"{_fmt(d['precio_bc3'], 4):>14}  "
                f"{_fmt(d['precio_calculado'], 4):>14}  "
                f"{_fmt(d['diferencia'], 2):>10}  "
                f"{d['diferencia_pct']:>6.2f}%  "
                f"{(d['resumen'] or '')[:40]}"
            )
        if len(val["precios"]) > 30:
            _safe_print(f"     ... y {len(val['precios']) - 30} mas")

    # ---- 3) PEM -----------------------------------------------------------
    _safe_print()
    _safe_print(f"  [3] PEM TOTAL")
    _safe_print(subsep)
    if not val["pem"]:
        _safe_print("     OK (o sin precios congelados en capitulos raiz que comparar).")
    else:
        pem = val["pem"]
        _safe_print(f"     PEM declarado en el archivo (suma capitulos raiz): "
                    f"{_fmt(pem['pem_archivo'])} EUR")
        _safe_print(f"     PEM propagado desde partidas:                     "
                    f"{_fmt(pem['pem_calculado'])} EUR")
        _safe_print(f"     Diferencia: {_fmt(pem['diferencia'])} EUR")

    _safe_print(sep)
    _safe_print()


# ============ Historial Undo/Redo ============
# Modelo estándar: `undo_stack` es una lista cuyo ÚLTIMO elemento es SIEMPRE el
# estado actual. `redo_stack` guarda los estados deshechos.
#   - Cargar archivo     → undo_stack=[S0], redo_stack=[]
#   - Cada edición       → push del nuevo estado; se limpia redo_stack
#   - Undo               → mueve la cima a redo_stack y restaura la nueva cima
#   - Redo               → recupera de redo_stack y restaura
_MAX_HISTORIAL = 50   # nº máximo de estados guardados (además del inicial)

def _snapshot_texto():
    """Devuelve el BC3 del estado actual como texto, o None si no hay obra."""
    from bc3manager.io.escritor import EscritorBC3
    p = _estado.get("presupuesto")
    if not p:
        return None
    try:
        return EscritorBC3(p).generar_texto()
    except Exception:
        return None

def _init_historial():
    """Reinicia el historial con el estado actual como punto de partida."""
    snap = _snapshot_texto()
    _estado["undo_stack"] = [snap] if snap is not None else []
    _estado["redo_stack"] = []

def _restaurar_snapshot(snapshot: str):
    """Carga un snapshot (texto BC3) como presupuesto actual y lo guarda en disco."""
    import tempfile, os
    with tempfile.NamedTemporaryFile(mode="w", suffix=".bc3",
                                     encoding="cp1252", errors="replace",
                                     delete=False, newline="") as tmp:
        tmp.write(snapshot)
        ruta_tmp = tmp.name
    try:
        p = leer_bc3(ruta_tmp)
    finally:
        os.unlink(ruta_tmp)
    _estado["presupuesto"] = p
    ruta = _estado.get("ruta_original")
    if ruta:
        try:
            escribir_bc3(p, ruta)
        except Exception:
            pass
    return p

def _autoguardar():
    """Tras una edición: guarda en disco, apila el nuevo estado y limpia el redo."""
    p = _estado.get("presupuesto")
    if not p:
        return
    # Apilar el nuevo estado como cima del historial
    snap = _snapshot_texto()
    if snap is not None:
        pila = _estado.setdefault("undo_stack", [])
        # Si la pila está vacía (no se inicializó al cargar), siembra el estado actual.
        pila.append(snap)
        # Limitar profundidad (conservamos el inicial + _MAX_HISTORIAL ediciones)
        if len(pila) > _MAX_HISTORIAL + 1:
            del pila[0:len(pila) - (_MAX_HISTORIAL + 1)]
        # Cualquier edición nueva invalida la rama de rehacer
        _estado["redo_stack"] = []
    # Guardar en disco
    ruta = _estado.get("ruta_original")
    if ruta:
        try:
            escribir_bc3(p, ruta)
        except Exception:
            pass

@app.route("/api/undo", methods=["POST"])
def api_undo():
    """Deshace la última edición: restaura el estado anterior del historial."""
    undo = _estado.get("undo_stack") or []
    if len(undo) < 2:
        return jsonify({"error": "No hay nada que deshacer"}), 400
    try:
        actual = undo.pop()                       # quita el estado actual
        _estado.setdefault("redo_stack", []).append(actual)
        snapshot = undo[-1]                        # nueva cima = estado a restaurar
        _restaurar_snapshot(snapshot)
        return _resp()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/redo", methods=["POST"])
def api_redo():
    """Rehace la última acción deshecha."""
    redo = _estado.get("redo_stack") or []
    if not redo:
        return jsonify({"error": "No hay nada que rehacer"}), 400
    try:
        snapshot = redo.pop()
        _estado.setdefault("undo_stack", []).append(snapshot)
        _restaurar_snapshot(snapshot)
        return _resp()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _generar_excel_validacion(p, ruta_destino: str) -> str:
    """Genera un .xlsx con la comparación completa archivo vs propagado.
    Hojas: Resumen, Importes partidas, Importes capitulos, Mediciones,
    Precios, PEM, Detalle conceptos. Devuelve la ruta del fichero generado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    val = p.validar_completo()
    comp = p.comparar_importes_archivo()
    wb = Workbook()

    # Estilos compartidos
    head_font = Font(bold=True, color="FFFFFF", size=11)
    head_fill = PatternFill("solid", fgColor="2B6FD0")
    ok_fill   = PatternFill("solid", fgColor="E8F5E9")
    bad_fill  = PatternFill("solid", fgColor="FFEBEE")
    bord = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def write_header(ws, cols):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bord

    def autosize(ws):
        for col in ws.columns:
            ml = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 60)

    # ---- Hoja 1: Resumen ----
    ws = wb.active
    ws.title = "Resumen"
    r = val["resumen"]
    n_partidas_diff = sum(1 for d in comp["partidas"] if abs(d["diferencia"]) >= 0.01)
    n_caps_diff = sum(1 for d in comp["capitulos"] if abs(d["diferencia"]) >= 0.01)
    rows = [
        ("Archivo", _estado.get("nombre_archivo", "")),
        ("Conceptos en el presupuesto", r["total_conceptos"]),
        ("Partidas analizadas", len(comp["partidas"])),
        ("Capitulos analizados", len(comp["capitulos"])),
        ("Mediciones (registros ~M)", r["total_mediciones"]),
        ("", ""),
        ("PEM declarado en el archivo (€)", comp["pem_archivo"]),
        ("PEM propagado desde partidas (€)", comp["pem_calculado"]),
        ("Diferencia PEM (€)",               comp["diferencia_pem"]),
        ("", ""),
        ("Partidas con diferencia archivo vs calc", n_partidas_diff),
        ("Capitulos con diferencia archivo vs calc", n_caps_diff),
        ("Mediciones con suma de lineas != total declarado", len(val["mediciones"])),
        ("Precios con archivo != recalculado", len(val["precios"])),
    ]
    for i, (k, v) in enumerate(rows, 1):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        c1.font = Font(bold=True)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            if "(€)" in k:
                c2.number_format = "#,##0.00"
                c2.alignment = Alignment(horizontal="right")
        if k.startswith("Diferencia PEM") or k.startswith("Partidas con") or k.startswith("Capitulos con") or k.startswith("Mediciones con") or k.startswith("Precios con"):
            es_bad = (isinstance(v, (int, float)) and abs(v) > 0.01)
            c2.fill = bad_fill if es_bad else ok_fill
    autosize(ws)

    # ---- Hoja 2: Importes partidas (TODAS, no solo discrepancias) ----
    ws = wb.create_sheet("Importes partidas")
    write_header(ws, [
        "Codigo", "En capitulo", "Descripcion", "Ud",
        "Precio archivo", "Precio calc",
        "Medicion",
        "Importe archivo", "Importe calc",
        "Diferencia", "% dif",
    ])
    for i, d in enumerate(comp["partidas"], 2):
        row = [
            d["codigo"], d["padre"], d["resumen"], d["unidad"],
            d["precio_archivo"], d["precio_calc"],
            d["medicion"],
            d["importe_archivo"], d["importe_calc"],
            d["diferencia"], d["diferencia_pct"],
        ]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = bord
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.0000" if j <= 7 else "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            if j == 10 and isinstance(v, (int, float)) and abs(v) >= 0.01:
                cell.fill = bad_fill
    # Fila total al final
    if comp["partidas"]:
        fila_total = len(comp["partidas"]) + 2
        tot_arch = sum(d["importe_archivo"] for d in comp["partidas"])
        tot_calc = sum(d["importe_calc"] for d in comp["partidas"])
        ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=fila_total, column=8, value=round(tot_arch, 2)).font = Font(bold=True)
        ws.cell(row=fila_total, column=9, value=round(tot_calc, 2)).font = Font(bold=True)
        ws.cell(row=fila_total, column=10, value=round(tot_calc - tot_arch, 2)).font = Font(bold=True)
        for j in (8, 9, 10):
            c = ws.cell(row=fila_total, column=j)
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
    autosize(ws)

    # ---- Hoja 3: Importes capitulos (TODOS) ----
    ws = wb.create_sheet("Importes capitulos")
    write_header(ws, [
        "Codigo", "En padre", "Descripcion",
        "Precio archivo (~C)",
        "Importe archivo", "Importe calc (suma hijos)",
        "Diferencia", "% dif",
    ])
    for i, d in enumerate(comp["capitulos"], 2):
        row = [
            d["codigo"], d["padre"], d["resumen"],
            d["precio_archivo"],
            d["importe_archivo"], d["importe_calc"],
            d["diferencia"], d["diferencia_pct"],
        ]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = bord
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.0000" if j <= 4 else "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            if j == 7 and isinstance(v, (int, float)) and abs(v) >= 0.01:
                cell.fill = bad_fill
    autosize(ws)

    # ---- Hoja 4: Mediciones ----
    ws = wb.create_sheet("Mediciones")
    write_header(ws, ["Partida", "En capitulo", "Descripcion",
                      "Total declarado (~M)", "Suma de lineas",
                      "Diferencia", "% diferencia"])
    for i, d in enumerate(val["mediciones"], 2):
        denom = max(abs(d["declarada"]), abs(d["suma_lineas"])) or 1
        pct = abs(d["diferencia"]) / denom * 100
        row = [d["codigo"], d["padre"], d["resumen"],
               d["declarada"], d["suma_lineas"], d["diferencia"], round(pct, 2)]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = bord
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.0000"
                cell.alignment = Alignment(horizontal="right")
    autosize(ws)

    # ---- Hoja 3: Precios ----
    ws = wb.create_sheet("Precios")
    write_header(ws, ["Codigo", "Tipo", "Descripcion",
                      "Precio archivo (~C)", "Precio calculado",
                      "Diferencia", "% diferencia"])
    for i, d in enumerate(val["precios"], 2):
        tipo = "Capitulo" if d["tipo"] == "capitulo" else "Partida"
        row = [d["codigo"], tipo, d["resumen"],
               d["precio_bc3"], d["precio_calculado"],
               d["diferencia"], d["diferencia_pct"]]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = bord
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.0000"
                cell.alignment = Alignment(horizontal="right")
    autosize(ws)

    # ---- Hoja 4: PEM ----
    ws = wb.create_sheet("PEM")
    if val["pem"]:
        pem = val["pem"]
        rows = [
            ("PEM declarado en archivo (€)",     pem["pem_archivo"]),
            ("PEM propagado desde partidas (€)", pem["pem_calculado"]),
            ("Diferencia (€)",                   pem["diferencia"]),
        ]
    else:
        rows = [
            ("Estado", "OK — sin discrepancias significativas a nivel PEM"),
            ("PEM (€)", r["pem_calculado"]),
        ]
    for i, (k, v) in enumerate(rows, 1):
        c1 = ws.cell(row=i, column=1, value=k); c1.font = Font(bold=True)
        c2 = ws.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)):
            c2.number_format = "#,##0.00"
            c2.alignment = Alignment(horizontal="right")
    autosize(ws)

    # ---- Hoja 5: Detalle conceptos ----
    # Volcado completo: para cada concepto, qué dice el archivo y qué el cálculo.
    ws = wb.create_sheet("Detalle conceptos")
    write_header(ws, ["Codigo", "Tipo", "Descripcion", "Unidad",
                      "Precio archivo", "Precio calculado",
                      "Suma mediciones", "Importe (precio_calc x med)"])
    fila = 2
    for cod, c in p.conceptos.items():
        pbc3 = getattr(c, "_precio_bc3", None)
        # Suma total de mediciones de este concepto en todos los padres
        suma_med = 0.0
        for m in c.mediciones.values():
            suma_med += sum(ln.subtotal for ln in m.lineas)
        importe = round(c.precio * suma_med, 2) if suma_med else 0.0
        row = [
            cod, c.tipo.value, c.resumen, c.unidad,
            pbc3 if pbc3 is not None else "", c.precio,
            round(suma_med, 4), importe,
        ]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=fila, column=j, value=v)
            cell.border = bord
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.0000"
                cell.alignment = Alignment(horizontal="right")
        fila += 1
    autosize(ws)

    wb.save(ruta_destino)
    return ruta_destino


def _ruta_validacion_xlsx() -> str:
    """Decide dónde guardar el Excel de validación: junto al BC3 si la ruta
    es real, o en temp si vino por upload web."""
    ruta_bc3 = _estado.get("ruta_original") or ""
    nombre_bc3 = _estado.get("nombre_archivo") or "presupuesto.bc3"
    base = os.path.splitext(os.path.basename(nombre_bc3))[0]
    nombre_xlsx = f"validacion_{base}.xlsx"
    if ruta_bc3:
        # ¿Está en temp? entonces guardar también en temp
        try:
            es_temporal = os.path.commonpath([ruta_bc3, tempfile.gettempdir()]) == tempfile.gettempdir()
        except (ValueError, OSError):
            es_temporal = False
        if es_temporal:
            return os.path.join(tempfile.gettempdir(), nombre_xlsx)
        # Junto al BC3 original
        return os.path.join(os.path.dirname(os.path.abspath(ruta_bc3)), nombre_xlsx)
    return os.path.join(tempfile.gettempdir(), nombre_xlsx)


def _generar_validacion_xlsx_seguro(p) -> Optional[str]:
    """Genera el Excel de validación y devuelve la ruta. Si falla, devuelve None
    sin reventar la carga del archivo."""
    try:
        destino = _ruta_validacion_xlsx()
        _generar_excel_validacion(p, destino)
        return destino
    except Exception as e:
        _safe_print(f"[AVISO] No se pudo generar el Excel de validacion: {e}")
        return None


@app.route("/api/validacion_xlsx")
def api_validacion_xlsx():
    """Descarga manual del Excel de validación (lo genera al vuelo)."""
    p = _estado.get("presupuesto")
    if not p: return "No hay presupuesto cargado", 400
    try:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); tmp.close()
        _generar_excel_validacion(p, tmp.name)
    except ImportError:
        return "Falta el modulo openpyxl. Instala: pip install openpyxl", 500
    nombre = (_estado.get("nombre_archivo") or "presupuesto.bc3").rsplit(".", 1)[0]
    return send_file(
        tmp.name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"validacion_{nombre}.xlsx",
    )

@app.route("/api/informe")
def api_informe():
    p = _estado.get("presupuesto")
    if not p: return "No hay presupuesto cargado", 400
    return Response(generar_informe(p, request.args.get("tipo", "presupuesto")), mimetype="text/html")

@app.route("/api/exportar")
def api_exportar():
    p = _estado.get("presupuesto")
    if not p: return "No hay presupuesto cargado", 400
    tmp = tempfile.NamedTemporaryFile(suffix=".bc3", delete=False); tmp.close()
    escribir_bc3(p, tmp.name)
    return send_file(tmp.name, mimetype="application/octet-stream", as_attachment=True, download_name="presupuesto_exportado.bc3")

@app.route("/api/editar", methods=["POST"])
def api_editar():
    """Endpoint único de edición. Recibe {accion, ...params}."""
    p = _estado.get("presupuesto")
    if not p: return jsonify({"error": "No hay presupuesto cargado"}), 400
    d = request.json
    accion = d.get("accion", "")
    try:
        if accion == "precio":
            p.modificar_precio(d["codigo"], float(d["valor"]))
        elif accion == "resumen":
            p.modificar_resumen(d["codigo"], d["valor"])
        elif accion == "unidad":
            p.modificar_unidad(d["codigo"], d["valor"])
        elif accion == "codigo":
            p.modificar_codigo(d["codigo_viejo"], d["codigo_nuevo"])
        elif accion == "rendimiento":
            p.modificar_rendimiento(d["codigo_padre"], d["codigo_hijo"], float(d["valor"]))
        elif accion == "medicion":
            if d["campo"] == "comentario":
                c = p.conceptos.get(d["codigo_hijo"])
                med = c.mediciones.get(d["codigo_padre"])
                med.lineas[int(d["indice"])].comentario = d["valor"]
            else:
                p.modificar_medicion(d["codigo_hijo"], d["codigo_padre"], int(d["indice"]), d["campo"], float(d["valor"]))
        elif accion == "add_linea_medicion":
            p.add_linea_medicion(d["codigo_hijo"], d["codigo_padre"], comentario=d.get("comentario",""), n_uds=float(d.get("n_uds",0)), longitud=float(d.get("longitud",0)), anchura=float(d.get("anchura",0)), altura=float(d.get("altura",0)))
        elif accion == "eliminar_linea_medicion":
            p.eliminar_linea_medicion(d["codigo_hijo"], d["codigo_padre"], int(d["indice"]))
        elif accion == "add_partida":
            p.add_partida(d["codigo_padre"], d["codigo"], d.get("unidad",""), d.get("resumen",""), float(d.get("precio",0)))
        elif accion == "add_capitulo":
            p.add_capitulo(d["codigo"], d.get("resumen",""), d.get("codigo_padre"))
        elif accion == "eliminar_concepto":
            p.eliminar_concepto(d["codigo"], d["codigo_padre"])
        elif accion == "add_recurso":
            p.add_recurso(d["codigo_partida"], d["codigo_recurso"],
                          float(d.get("rendimiento", 1)),
                          float(d.get("precio", 0)),
                          d.get("unidad", ""), d.get("resumen", ""))
        elif accion == "eliminar_recurso":
            p.eliminar_recurso(d["codigo_partida"], d["codigo_recurso"])
        elif accion == "reordenar_recurso":
            p.reordenar_recurso(d["codigo_partida"], d["codigo_recurso"], d.get("antes_de"))
        elif accion == "reordenar_medicion":
            p.reordenar_medicion(d["codigo_hijo"], d["codigo_padre"], int(d["from_idx"]), int(d["to_idx"]))
        elif accion == "texto":
            p.modificar_texto(d["codigo"], d["valor"])
        elif accion == "cambiar_tipo":
            p.cambiar_tipo(d["codigo"], d["tipo"])
        elif accion == "tipo_recurso":
            p.cambiar_tipo_recurso(d["codigo"], d["tipo_fiebdc"])
        elif accion == "mover":
            p.mover_concepto(d["codigo"], d["padre_origen"], d["padre_destino"],
                             d.get("antes_de"))   # antes_de puede ser None → append
        elif accion == "copiar":
            p.copiar_concepto(d["codigo"], d["padre_destino"], d.get("antes_de"))
        else:
            return jsonify({"error": f"Acción desconocida: {accion}"}), 400
        # Ya no es necesario invalidar precio_es_dato — recalcular() suma siempre
        # desde los descendientes (los precios de capítulo del archivo se ignoran).
        _autoguardar()
        return _resp()
    except Exception as e:
        return jsonify({"error": str(e)}), 400
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>BC3Manager</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/tabulator-tables@6.3.0/dist/css/tabulator.min.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/tabulator-tables@6.3.0/dist/js/tabulator.min.js"></script>
<style>
:root,html[data-theme=dark]{--bg:#0f1117;--bg-card:#1a1d27;--bg-hover:#222635;--bg-active:#2a2f3f;--border:#2e3346;--border-light:#383d52;--text:#e4e6ed;--text-dim:#8b90a5;--text-muted:#5c6178;--accent:#4f8ff7;--accent-soft:rgba(79,143,247,.12);--accent-hover:#6ba1f9;--green:#3ecf8e;--green-soft:rgba(62,207,142,.12);--amber:#f0b429;--red:#ef6b6b;--radius:8px;--radius-lg:12px;--shadow:0 2px 12px rgba(0,0,0,.3)}
html[data-theme=light]{--bg:#f5f6f8;--bg-card:#ffffff;--bg-hover:#eef0f4;--bg-active:#e4e7ee;--border:#d8dbe3;--border-light:#c8ccd6;--text:#1a1d27;--text-dim:#555a6e;--text-muted:#888da1;--accent:#2b6fd0;--accent-soft:rgba(43,111,208,.1);--accent-hover:#3b7fe0;--green:#1a9960;--green-soft:rgba(26,153,96,.1);--amber:#c8900a;--red:#d04040;--shadow:0 2px 8px rgba(0,0,0,.08)}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'DM Sans',system-ui,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
button{cursor:pointer;font-family:inherit}
.header{background:var(--bg-card);border-bottom:1px solid var(--border);padding:0 32px;height:56px;display:flex;align-items:center;gap:16px;position:sticky;top:0;z-index:100}
.header .logo{font-size:15px;font-weight:700;letter-spacing:-.5px;color:var(--accent)}.header .logo span{color:var(--text-dim);font-weight:400}
.header .sep{width:1px;height:24px;background:var(--border)}
.btn{display:inline-flex;align-items:center;gap:6px;padding:7px 14px;border-radius:var(--radius);border:1px solid var(--border);background:var(--bg-card);color:var(--text-dim);font-size:13px;font-weight:500;transition:all .15s}
.btn:hover{background:var(--bg-hover);color:var(--text);border-color:var(--border-light)}
.btn-accent{background:var(--accent);color:#fff;border-color:var(--accent)}.btn-accent:hover{background:var(--accent-hover);border-color:var(--accent-hover)}
.btn-sm{padding:3px 8px;font-size:11px}
.btn-danger{border-color:var(--red);color:var(--red)}
.btn svg{width:15px;height:15px}
.header-right{margin-left:auto;display:flex;gap:8px;align-items:center}
.stats-bar{background:var(--bg-card);border-bottom:1px solid var(--border);padding:12px 32px;display:flex;gap:32px;flex-wrap:wrap}
.stat{display:flex;flex-direction:column;gap:2px}.stat-label{font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted)}.stat-value{font-size:14px;font-weight:600}.stat-value.total{color:var(--green);font-size:18px}
.upload-screen{display:flex;align-items:center;justify-content:center;min-height:calc(100vh - 56px)}
.upload-box{border:2px dashed var(--border);border-radius:var(--radius-lg);padding:64px 80px;text-align:center;transition:all .2s;cursor:pointer}
.upload-box:hover,.upload-box.drag{border-color:var(--accent);background:var(--accent-soft)}
.upload-box h2{font-size:20px;font-weight:600;margin-bottom:8px}.upload-box p{color:var(--text-dim);font-size:14px;margin-bottom:24px}
.main{display:flex;height:calc(100vh - 56px - 49px);overflow:hidden}
.tree-panel{width:60%;min-width:520px;border-right:1px solid var(--border);display:flex;flex-direction:column;overflow:hidden;background:var(--bg)}
.detail-panel{flex:1;overflow-y:auto;padding:20px 24px;min-width:380px}
.tree-scroll{overflow:auto;flex:1}
/* Tabla del árbol */
/* min-width garantiza que la columna Descripción no se colapse cuando el panel
   es estrecho. Las columnas fijas suman ~448px; reservamos ~260px para la
   descripción. Si no cabe, el contenedor (.tree-scroll) hace scroll. */
.ttable{width:100%;border-collapse:collapse;font-size:12px;table-layout:fixed;min-width:710px}
.ttable thead{position:sticky;top:0;z-index:5}
.ttable thead th{background:var(--bg-card);color:var(--text-dim);text-align:left;padding:9px 8px;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.4px;border-bottom:1px solid var(--border);user-select:none}
.ttable thead th.num{text-align:right}
.ttable tbody td{padding:5px 8px;border-bottom:1px solid var(--border);vertical-align:middle;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.ttable tbody tr{cursor:pointer;transition:background .08s}
.ttable tbody tr:hover td{background:var(--bg-hover)}
.ttable tbody tr.active td{background:var(--bg-active);box-shadow:inset 3px 0 0 var(--accent)}
.ttable .num{text-align:right;font-family:'JetBrains Mono',monospace;font-size:11px;font-variant-numeric:tabular-nums}
.ttable .cod{font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--text-dim)}
.ttable .imp{color:var(--green);font-weight:500}
/* Columnas (anchos) */
.col-cod{width:112px}
.col-ud{width:46px}
.col-cant{width:72px}
.col-precio{width:72px}
.col-imp{width:92px}
.col-act{width:54px}
/* Fila por tipo */
.ttable tr.row-cap td{background:var(--bg-card)}
.ttable tr.row-cap td.col-resumen{font-weight:600}
.ttable tr.row-cap:hover td{background:var(--bg-hover)}
.ttable tr.row-cap.active td{background:var(--bg-active)}
/* Toggle de despliegue */
.tt-toggle{display:inline-flex;width:14px;height:14px;align-items:center;justify-content:center;color:var(--text-muted);font-size:10px;margin-right:4px;transition:transform .15s;user-select:none;vertical-align:middle}
.tt-toggle.open{transform:rotate(90deg)}
.tt-toggle.leaf{visibility:hidden}
.tt-indent{display:inline-block;vertical-align:middle}
.tt-badge{display:inline-block;font-size:9px;font-weight:700;letter-spacing:.4px;padding:1px 5px;border-radius:3px;margin-right:6px;vertical-align:middle}
.badge-cap{background:var(--accent-soft);color:var(--accent)}
.badge-part{background:var(--green-soft);color:var(--green)}
/* Botones de acción por fila (aparecen al hover) */
.row-actions{display:inline-flex;gap:3px;opacity:0;transition:opacity .1s;justify-content:flex-end;width:100%}
.ttable tbody tr:hover .row-actions,.ttable tbody tr.active .row-actions{opacity:1}
.row-actions button{width:22px;height:22px;padding:0;border-radius:4px;border:1px solid var(--border);background:var(--bg-card);color:var(--text-dim);font-size:13px;line-height:1;display:inline-flex;align-items:center;justify-content:center;cursor:pointer;transition:all .1s}
.row-actions button:hover{color:var(--text);background:var(--bg-hover);border-color:var(--border-light)}
.row-actions button.danger:hover{color:var(--red);border-color:var(--red)}
/* Filas fantasma para añadir */
.ghost-row td{font-size:12px;padding:5px 8px;border-bottom:1px solid var(--border);cursor:default;user-select:none}
.ghost-row:hover td{background:var(--accent-soft)}
.ghost-row .ecell{min-width:70px;min-height:18px;border:1px dashed var(--border-light);border-radius:3px;display:inline-block;background:var(--bg-card);vertical-align:middle}
.ghost-row .ecell.num{min-width:50px}
.ghost-row td.col-cod .ecell:empty:before{content:'código\2026';color:var(--text-muted);font-size:10px;font-style:italic;pointer-events:none}
/* Drag & drop */
.drag-handle{cursor:grab;color:var(--text-muted);font-size:13px;padding:1px 4px;user-select:none;opacity:0;transition:opacity .1s;line-height:1}
.ttable tbody tr:hover .drag-handle,.ttable tbody tr.active .drag-handle{opacity:.45}
.drag-handle:hover{opacity:.9!important}
.drag-handle:active{cursor:grabbing}
.ttable tr.dragging td{opacity:.3}
.ttable tr.drop-before td{box-shadow:inset 0 2px 0 var(--accent)}
.ttable tr.drop-after td{box-shadow:inset 0 -2px 0 var(--accent)}
.ttable tr.drop-into td{background:var(--accent-soft)!important;box-shadow:inset 2px 0 0 var(--accent)}
.detail-empty{display:flex;align-items:center;justify-content:center;height:100%;color:var(--text-muted);font-size:14px}
.detail-header{margin-bottom:16px;padding-bottom:16px;border-bottom:1px solid var(--border)}
.detail-code{font-family:'JetBrains Mono',monospace;font-size:12px;color:var(--accent);margin-bottom:4px}
.detail-name{font-size:20px;font-weight:700;margin-bottom:8px}
.detail-meta{display:flex;gap:24px;flex-wrap:wrap}
.detail-meta-item{display:flex;flex-direction:column;gap:2px}
.detail-meta-label{font-size:11px;color:var(--text-muted);text-transform:uppercase;letter-spacing:.5px}
.detail-meta-value{font-size:15px;font-weight:600}.detail-meta-value.green{color:var(--green)}
.detail-section{margin-top:20px}.detail-section h3{font-size:13px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;color:var(--text-dim);margin-bottom:10px}
.detail-text{font-size:13px;line-height:1.6;color:var(--text-dim);background:var(--bg-card);padding:12px 16px;border-radius:var(--radius);border:1px solid var(--border);white-space:pre-wrap;word-break:break-word;min-height:48px;cursor:text;transition:border-color .15s,outline .15s}
.detail-text:focus{outline:2px solid var(--accent);outline-offset:-1px;border-color:var(--accent);color:var(--text)}
.detail-text:empty:before{content:attr(data-placeholder);color:var(--text-muted);font-style:italic;pointer-events:none}
/* Selector de tipo (Cap/Part en árbol; MO/MQ/MT/AUX en desglose) */
.tt-badge.clickable{cursor:pointer}
.tt-badge.clickable:hover{opacity:.75}
.tipo-dropdown{position:absolute;z-index:200;background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius);box-shadow:var(--shadow);padding:4px;min-width:90px}
.tipo-dropdown-item{padding:5px 10px;cursor:pointer;border-radius:4px;font-size:12px;white-space:nowrap}
.tipo-dropdown-item:hover{background:var(--bg-hover)}
.tipo-dropdown-item.active{font-weight:700;color:var(--accent)}
/* Colores por subtipo de recurso */
.badge-mo{background:rgba(240,180,41,.18);color:var(--amber)}
.badge-mq{background:var(--accent-soft);color:var(--accent)}
.badge-mt{background:var(--green-soft);color:var(--green)}
.badge-aux{background:rgba(180,120,220,.15);color:#b48adf}
.dtable{width:100%;border-collapse:collapse;font-size:12px;border:1px solid var(--border);border-radius:var(--radius);overflow:hidden}
.dtable th{background:var(--bg-card);text-align:left;padding:8px 10px;font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.3px;color:var(--text-dim);border-bottom:1px solid var(--border)}
.dtable td{padding:5px 10px;border-bottom:1px solid var(--border);vertical-align:middle}
.dtable tr:last-child td{border-bottom:none}.dtable tr:hover td{background:var(--bg-hover)}
.dtable .num{text-align:right;font-family:'JetBrains Mono',monospace;font-size:11px}
.dtable .total-row td{font-weight:700;background:var(--bg-card);color:var(--green)}
/* editable cells */
.ecell{cursor:default;border-radius:3px;padding:2px 4px;transition:background .1s;min-width:30px;min-height:18px;display:inline-block;user-select:none;vertical-align:middle}
.ecell[data-editable=true]:hover{background:var(--bg-active);cursor:default}
.ecell[data-editable=true]:focus{outline:1px dashed var(--border-light);outline-offset:-1px}
.ecell[contenteditable=true]{cursor:text;user-select:text;outline:2px solid var(--accent) !important;background:var(--bg) !important;outline-offset:-1px}
/* Celda editable vacía: muestra borde punteado para que sea visible y clicable */
.ecell[data-editable=true]:empty{border:1px dashed var(--border-light);background:var(--bg-card)}
.actions-bar{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}
.dropdown{position:relative}
.dropdown-menu{display:none;position:absolute;top:100%;right:0;margin-top:4px;background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius);box-shadow:var(--shadow);min-width:200px;z-index:50;padding:4px}
.dropdown-menu.open{display:block}
.dropdown-item{display:block;width:100%;padding:8px 12px;text-align:left;font-size:13px;color:var(--text);background:none;border:none;border-radius:6px;transition:background .1s}
.dropdown-item:hover{background:var(--bg-hover)}
/* Selección múltiple en árbol */
.ttable tr.tree-selected td{background:var(--accent-soft)!important;box-shadow:inset 2px 0 0 var(--accent)}
.ttable tr.tree-selected:hover td{background:var(--bg-active)!important}
/* Fila copiada al portapapeles */
.ttable tr.copied-row td{outline:1px dashed var(--accent);outline-offset:-1px}
/* Toast de notificación */
#_toast{position:fixed;bottom:24px;left:50%;transform:translateX(-50%);background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius);padding:8px 18px;font-size:12px;color:var(--text);box-shadow:var(--shadow);z-index:300;pointer-events:none;transition:opacity .35s;white-space:nowrap}
.loading{display:flex;align-items:center;justify-content:center;height:100%;gap:12px;color:var(--text-dim)}
.spinner{width:20px;height:20px;border:2px solid var(--border);border-top-color:var(--accent);border-radius:50%;animation:spin .6s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
@media(max-width:768px){.main{flex-direction:column}.tree-panel{width:100%;height:50vh;border-right:none;border-bottom:1px solid var(--border)}.header{padding:0 16px}.detail-panel{padding:16px}}
/* ===== Tabulator — tema BC3Manager ===== */
.tabulator{background:transparent!important;border:1px solid var(--border)!important;border-radius:var(--radius)!important;font-size:12px!important;font-family:'DM Sans',system-ui,sans-serif!important;color:var(--text)!important}
.tabulator .tabulator-header{background:var(--bg-card)!important;border-bottom:1px solid var(--border)!important}
.tabulator .tabulator-header .tabulator-col{background:var(--bg-card)!important;border-right:1px solid var(--border)!important;color:var(--text-dim)!important;font-size:11px!important;font-weight:600!important;text-transform:uppercase;letter-spacing:.3px!important;padding:8px 10px!important}
.tabulator .tabulator-header .tabulator-col:last-child{border-right:none!important}
.tabulator .tabulator-tableholder{background:transparent!important}
.tabulator-row{background:transparent!important;border-bottom:1px solid var(--border)!important;color:var(--text)!important}
.tabulator-row:hover{background:var(--bg-hover)!important}
.tabulator-row.tabulator-selected,.tabulator-row.tabulator-selected:hover{background:var(--accent-soft)!important}
.tabulator-row.tabulator-moving{opacity:.5!important}
.tabulator-row .tabulator-cell{border-right:1px solid var(--border)!important;padding:5px 10px!important;color:var(--text)!important;overflow:hidden;text-overflow:ellipsis}
.tabulator-row .tabulator-cell:last-child{border-right:none!important}
.tabulator-row .tabulator-cell.tabulator-editing{outline:2px solid var(--accent)!important;outline-offset:-2px!important;background:var(--bg)!important;padding:0!important}
.tabulator-row .tabulator-cell.tabulator-editing input,.tabulator-row .tabulator-cell.tabulator-editing select{background:var(--bg)!important;color:var(--text)!important;border:none!important;outline:none!important;font-family:inherit!important;font-size:12px!important;padding:5px 10px!important;width:100%!important;height:100%!important}
.tabulator-row .tabulator-cell .tabulator-data-tree-control{color:var(--text-muted)!important}
.tabulator .tabulator-footer{background:var(--bg-card)!important;border-top:1px solid var(--border)!important;color:var(--text-dim)!important;font-weight:700!important;padding:5px 10px!important;font-size:12px!important}
.tabulator-placeholder{color:var(--text-muted)!important;font-style:italic;padding:16px!important}
.tabulator .tabulator-move-handle{color:var(--text-muted)!important}
.tab-num{text-align:right!important;font-family:'JetBrains Mono',monospace!important;font-size:11px!important}
.tab-actions{display:flex;align-items:center;gap:18px;margin-top:8px}
.tab-add-row{display:inline-flex;align-items:center;gap:6px;color:var(--text-muted);font-size:12px;cursor:pointer;padding:4px 2px;border-radius:var(--radius);transition:color .1s}
.tab-add-row:hover{color:var(--accent)}
.tab-del-row{display:inline-flex;align-items:center;gap:6px;color:var(--text-muted);font-size:12px;cursor:pointer;padding:4px 2px;border-radius:var(--radius);transition:color .1s}
.tab-del-row:hover{color:var(--red)}
</style></head><body>
<header class="header">
<div class="logo">BC3<span>Manager</span></div><div class="sep"></div><span id="fileName" style="font-size:12px;color:var(--text-muted);max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"></span>
<button class="btn" onclick="document.getElementById('fileInput').click()"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>Abrir BC3</button>
<input type="file" id="fileInput" accept=".bc3,.BC3" style="display:none" onchange="uploadFile(this)">
<div class="header-right" id="headerActions" style="display:none">
<div class="dropdown"><button class="btn" onclick="toggleDropdown(this)">Informes</button>
<div class="dropdown-menu">
<button class="dropdown-item" onclick="descargarInforme('mediciones')">Mediciones</button>
<button class="dropdown-item" onclick="descargarInforme('cuadro')">Cuadro de precios</button>
<button class="dropdown-item" onclick="descargarInforme('presupuesto')">Presupuesto</button>
<button class="dropdown-item" onclick="descargarInforme('resumen')">Resumen</button>
</div></div>
<button class="btn" id="undoBtn" onclick="undoAction()" title="Deshacer (Ctrl+Z)" style="display:none">↩ Deshacer</button>
<button class="btn" id="redoBtn" onclick="redoAction()" title="Rehacer (Ctrl+Y)" style="display:none">↪ Rehacer</button>
<button class="btn" onclick="exportarBC3()">Exportar BC3</button>
<button class="btn" onclick="addCapitulo('')">+ Capítulo</button>
<button class="btn" id="themeBtn" onclick="toggleTheme()" title="Cambiar tema">🌙</button>
</div></header>
<div id="uploadScreen" class="upload-screen"><div class="upload-box" id="uploadBox" onclick="document.getElementById('fileInput').click()" ondragover="event.preventDefault();this.classList.add('drag')" ondragleave="this.classList.remove('drag')" ondrop="event.preventDefault();this.classList.remove('drag');handleDrop(event)"><h2>Abre un archivo BC3</h2><p>Arrastra aquí o haz clic</p><button class="btn btn-accent">Seleccionar archivo</button></div></div>
<div id="mainApp" style="display:none">
<div id="tempBanner" style="display:none;background:rgba(240,180,41,.12);border-bottom:1px solid var(--amber);padding:8px 32px;font-size:12px;color:var(--amber)">⚠ Este archivo está en una carpeta temporal. Tus cambios se guardan ahí, pero el archivo puede borrarse al reiniciar el equipo. Exporta a BC3 para conservarlo.</div>
<div id="discBanner" style="display:none;background:rgba(239,107,107,.12);border-bottom:1px solid var(--red);padding:8px 32px;font-size:12px;color:var(--red);display:flex;align-items:center;gap:12px">
  <span id="discBannerText">⚠ Precios del archivo no coinciden con el recálculo</span>
  <button class="btn btn-sm" onclick="mostrarDiscrepancias()" style="margin-left:auto">Ver detalle</button>
  <button class="btn btn-sm" onclick="window.open('/api/validacion_xlsx','_blank')" title="Descargar Excel completo">📊 Excel</button>
  <button class="btn btn-sm" onclick="document.getElementById('discBanner').style.display='none'" title="Ocultar">×</button>
</div>
<div class="stats-bar" id="statsBar"></div>
<div class="main">
  <div class="tree-panel">
    <div class="tree-scroll" id="treeContainer">
      <table class="ttable"><thead><tr>
        <th class="col-cod">Código</th>
        <th class="col-resumen">Descripción</th>
        <th class="col-ud">Ud</th>
        <th class="num col-cant">Cantidad</th>
        <th class="num col-precio">Precio</th>
        <th class="num col-imp">Importe</th>
        <th class="col-act"></th>
      </tr></thead><tbody id="treeBody"></tbody></table>
    </div>
  </div>
  <div class="detail-panel" id="detailPanel"><div class="detail-empty">Selecciona una partida</div></div>
</div></div>
<div id="loadingOverlay" style="display:none;position:fixed;inset:0;background:rgba(15,17,23,.85);z-index:200"><div class="loading" style="height:100%"><div class="spinner"></div><span>Leyendo BC3...</span></div></div>
<div id="discModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:250;align-items:center;justify-content:center">
  <div style="background:var(--bg-card);border:1px solid var(--border);border-radius:var(--radius-lg);max-width:900px;width:90%;max-height:80vh;display:flex;flex-direction:column;box-shadow:var(--shadow)">
    <div style="padding:16px 20px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:12px">
      <strong style="font-size:14px;color:var(--red)">Discrepancias de precios</strong>
      <span id="discCount" style="font-size:12px;color:var(--text-dim)"></span>
      <button class="btn btn-sm" onclick="window.open('/api/validacion_xlsx','_blank')" style="margin-left:auto" title="Exporta hoja Excel con resumen, mediciones, precios, PEM y detalle por concepto">📊 Exportar Excel</button>
      <button class="btn btn-sm" onclick="document.getElementById('discModal').style.display='none'">Cerrar</button>
    </div>
    <div style="padding:8px 20px;font-size:12px;color:var(--text-dim);border-bottom:1px solid var(--border);line-height:1.5">
      Los conceptos siguientes tienen un precio en el archivo BC3 que <u>no coincide</u> con el resultado de sumar sus componentes.
      Causas habituales: precios redondeados en el archivo, descomposiciones incompletas, o el archivo se generó con precios "congelados" (Presto 8.8).
      <br><strong>Capítulos</strong>: conservan el precio del archivo hasta que edites algo.
      <strong>Partidas</strong>: siempre muestran el precio calculado desde sus recursos (el del archivo se considera obsoleto).
    </div>
    <div id="discList" style="overflow:auto;flex:1;padding:0 20px 16px"></div>
  </div>
</div>

<script>
let treeData=[], fileInfo={}, curNode=null, curParent='';
let _clipboard=null;        // {codigo,resumen} o [{codigo,padre},...] para selección múltiple
let _discrepancias=[];      // Discrepancias entre precio BC3 y precio recalculado al cargar
let _pemValidacion=null;    // {archivo, calculado, diferencia} comparación PEM al cargar
let _ctxCopia='arbol';      // zona con el foco para Ctrl+C: 'arbol' | 'descomp' | 'medic'
let _tabDescomp=null;       // instancia Tabulator — desglose
let _tabMedic=null;         // instancia Tabulator — mediciones
let selectedNodes=[];       // [{codigo,padre}] — selección múltiple en árbol
let _selectAnchor=null;     // {codigo,padre} — ancla para shift+clic

// ---- Upload ----
function uploadFile(input){const f=input.files[0];if(f)sendFile(f)}
function handleDrop(e){const f=e.dataTransfer.files[0];if(f)sendFile(f)}
function sendFile(file){
  document.getElementById('loadingOverlay').style.display='';
  const fd=new FormData();fd.append('archivo',file);
  fetch('/api/cargar',{method:'POST',body:fd}).then(r=>r.json()).then(data=>{
    document.getElementById('loadingOverlay').style.display='none';
    if(data.error){alert('Error: '+data.error);return}
    fileInfo=data.info;treeData=data.arbol;
    _discrepancias=data.discrepancias||[];
    _pemValidacion=data.pem||null;
    if(data.archivo)document.getElementById('fileName').textContent=data.archivo;
    renderApp();
    actualizarBannerDiscrepancias();
  }).catch(err=>{document.getElementById('loadingOverlay').style.display='none';alert('Error: '+err)})
}

// Formatea un número a formato europeo con 2 decimales
function _fmtEur(n){
  return Number(n).toLocaleString('es-ES',{minimumFractionDigits:2,maximumFractionDigits:2});
}

// Muestra/oculta el banner. Salta si hay discrepancias de precio O diferencia de PEM.
function actualizarBannerDiscrepancias(){
  const b=document.getElementById('discBanner');
  const t=document.getElementById('discBannerText');
  if(!b)return;
  const nPrecio=_discrepancias?_discrepancias.length:0;
  const difPem=_pemValidacion?Math.abs(_pemValidacion.diferencia):0;
  const hayPem=difPem>=0.01;
  if(nPrecio>0 || hayPem){
    let msg='⚠ ';
    const partes=[];
    if(hayPem){
      const d=_pemValidacion.diferencia;
      partes.push(`PEM archivo ${_fmtEur(_pemValidacion.archivo)} € vs calculado ${_fmtEur(_pemValidacion.calculado)} € (${d>=0?'+':''}${_fmtEur(d)} €)`);
    }
    if(nPrecio>0){
      partes.push(`${nPrecio} concepto${nPrecio>1?'s':''} con precio inconsistente`);
    }
    t.textContent=msg+partes.join('  ·  ');
    b.style.display='flex';
  }else{
    b.style.display='none';
  }
}

function mostrarDiscrepancias(){
  const m=document.getElementById('discModal');
  const c=document.getElementById('discCount');
  const l=document.getElementById('discList');
  c.textContent=`${_discrepancias.length} concepto${_discrepancias.length===1?'':'s'}`;
  // Bloque PEM destacado arriba
  let pemHtml='';
  if(_pemValidacion){
    const d=_pemValidacion.diferencia;
    const cuadra=Math.abs(d)<0.01;
    pemHtml=`<div style="margin-top:12px;padding:12px 16px;border-radius:var(--radius);
      background:${cuadra?'var(--green-soft)':'rgba(239,107,107,.12)'};
      border:1px solid ${cuadra?'var(--green)':'var(--red)'}">
      <strong>PEM total</strong><br>
      Archivo: <b>${_fmtEur(_pemValidacion.archivo)} €</b> &nbsp;·&nbsp;
      Calculado: <b>${_fmtEur(_pemValidacion.calculado)} €</b> &nbsp;·&nbsp;
      Diferencia: <b style="color:${cuadra?'var(--green)':'var(--red)'}">${d>=0?'+':''}${_fmtEur(d)} €</b>
      ${cuadra?' ✓':''}
    </div>`;
  }
  l.innerHTML=pemHtml+`<table class="dtable" style="margin-top:12px">
    <tr><th>Código</th><th>Tipo</th><th>Descripción</th>
        <th class="num">Precio archivo</th>
        <th class="num">Precio calculado</th>
        <th class="num">Diferencia</th>
        <th class="num">%</th></tr>
    ${_discrepancias.map(d=>`<tr>
      <td style="font-family:'JetBrains Mono',monospace;font-size:11px">${esc(d.codigo)}</td>
      <td><span class="tt-badge ${d.tipo==='capitulo'?'badge-cap':'badge-part'}">${d.tipo==='capitulo'?'Cap':'Part'}</span></td>
      <td>${esc(d.resumen)}</td>
      <td class="num">${esc(d.precio_bc3.toString().replace('.',','))}</td>
      <td class="num">${esc(d.precio_calculado.toString().replace('.',','))}</td>
      <td class="num" style="color:${d.diferencia>=0?'var(--green)':'var(--red)'}">${(d.diferencia>=0?'+':'')+esc(d.diferencia.toString().replace('.',','))}</td>
      <td class="num">${esc(d.diferencia_pct.toString().replace('.',','))}%</td>
    </tr>`).join('')}
  </table>`;
  m.style.display='flex';
}

// ---- API ----
async function api(data){
  const r=await fetch('/api/editar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
  const j=await r.json();if(j.error){alert(j.error);return null}return j
}
// Actualiza la visibilidad de los botones Deshacer/Rehacer según la respuesta
function _actualizarUndoRedo(j){
  const u=document.getElementById('undoBtn');
  if(u)u.style.display=(j.undo_disponible?'':'none');
  const r=document.getElementById('redoBtn');
  if(r)r.style.display=(j.redo_disponible?'':'none');
}
function refresh(j){
  if(!j)return;
  // Capturar la celda con foco ANTES de destruir el DOM (Tab/Enter habrán ya movido el foco a la siguiente celda)
  const savedFocus=_captureFocusPos();
  fileInfo=j.info;treeData=j.arbol;
  _actualizarUndoRedo(j);
  renderStats();renderTree();
  if(curNode){const f=findNode(treeData,curNode.codigo);if(f){curNode=f;renderDetail(f)}else{curNode=null;if(_pasteHandler){document.removeEventListener('paste',_pasteHandler);_pasteHandler=null}document.getElementById('detailPanel').innerHTML='<div class="detail-empty">Concepto eliminado</div>'}}
  // Restaurar foco en la celda equivalente del nuevo DOM
  if(savedFocus)_restoreFocusPos(savedFocus);
}

// Captura {rowKey, colIdx} del activeElement si es una celda editable
function _captureFocusPos(){
  const el=document.activeElement;
  if(!el||!el.classList||!el.classList.contains('ecell'))return null;
  const tr=el.closest('tr');
  if(!tr||!tr.dataset.rowKey)return null;
  const cells=[...tr.querySelectorAll('.ecell[data-editable=true]')];
  const ci=cells.indexOf(el);
  if(ci<0)return null;
  return {rowKey:tr.dataset.rowKey,colIdx:ci};
}
// Activa la celda equivalente tras un re-render
function _restoreFocusPos(pos){
  const tr=document.querySelector(`[data-row-key="${CSS.escape(pos.rowKey)}"]`);
  if(!tr)return;
  const cells=[...tr.querySelectorAll('.ecell[data-editable=true]')];
  const tgt=cells[pos.colIdx]||cells[cells.length-1];
  if(tgt)ecActivate(tgt);
}
function findNode(nodes,cod){for(const n of nodes){if(n.codigo===cod)return n;if(n.hijos){const f=findNode(n.hijos,cod);if(f)return f}}return null}

// ---- Tabulator lifecycle ----
function _destroyTabs(){
  if(_tabDescomp){try{_tabDescomp.destroy()}catch(e){}; _tabDescomp=null;}
  if(_tabMedic){try{_tabMedic.destroy()}catch(e){}; _tabMedic=null;}
}

// ---- Guardado silencioso (sin refresh) — campos que no afectan cálculos ----
async function silentSave(params){
  const r=await fetch('/api/editar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(params)});
  const j=await r.json();
  if(j&&j.error){alert(j.error);return null;}
  if(j){treeData=j.arbol;fileInfo=j.info;_actualizarUndoRedo(j);}
  return j;
}

// ---- Guardado con recálculo — actualiza importes sin destruir Tabulator ----
async function calcSave(params,preloaded){
  const j=preloaded||(await api(params));
  if(!j)return null;
  treeData=j.arbol;fileInfo=j.info;
  renderStats();renderTree();
  _actualizarUndoRedo(j);
  if(curNode){
    // Buscar nodo equivalente en el árbol nuevo (mismo código Y mismo padre)
    const nNew=_findNodeInParent(treeData,curNode.codigo,curParent);
    if(nNew){
      curNode=nNew;
      _updateDetailHeader(nNew);
      // replaceData (no updateData): sustituye TODAS las filas con datos frescos del servidor.
      // Necesario porque rendimiento/precio del recurso pueden cambiar y el importe se
      // recalcula en el servidor. updateData solo refresca campos específicos.
      if(_tabDescomp){
        const recs=(nNew.recursos||[]).map(r=>({...r}));
        try{_tabDescomp.replaceData(recs);}catch(e){console.error('replaceData descomp:',e);}
      }
      if(_tabMedic){
        const meds=(nNew.lineas_medicion||[]).map((ln,i)=>({...ln,_idx:i}));
        try{_tabMedic.replaceData(meds);}catch(e){console.error('replaceData medic:',e);}
      }
    }
  }
  return j;
}
// Busca un nodo por (codigo, padre) para evitar coger un homónimo de otro capítulo
function _findNodeInParent(nodes,cod,padre){
  for(const n of nodes){
    if(n.codigo===cod&&n.padre===padre)return n;
    if(n.hijos){const f=_findNodeInParent(n.hijos,cod,padre);if(f)return f;}
  }
  // Fallback: si no se encontró con el padre exacto, busca solo por código
  return findNode(nodes,cod);
}

// Renderiza el contenido interno de la celda Cantidad de la cabecera
// (editable si la partida no tiene mediciones, bloqueada y calculada si las tiene)
function _renderHeaderCant(nodo,cod,pc){
  return (nodo.lineas_medicion&&nodo.lineas_medicion.length>0)
    ? `<span class="ecell" contenteditable="false" style="cursor:default;opacity:.7" title="Calculada desde las mediciones">${esc(nodo.medicion_fmt||'0,00')}</span> <span style="font-size:11px;color:var(--text-muted)">🔒</span>`
    : `${ec(nodo.medicion_fmt||'0,00',true,v=>setCantidadSimple(cod,pc,parseNum(v)),true)}`;
}
// Renderiza el contenido interno de la celda Precio
function _renderHeaderPrecio(nodo,cod){
  return (nodo.recursos&&nodo.recursos.length>0)
    ? `<span class="ecell" contenteditable="false" style="cursor:default;opacity:.7" title="Calculado desde la descomposición">${esc(nodo.precio_fmt)}</span> € <span style="font-size:11px;color:var(--text-muted)">🔒</span>`
    : `${ec(nodo.precio_fmt,true,v=>api({accion:'precio',codigo:cod,valor:parseNum(v)}).then(refresh),true)} €`;
}
// Actualiza las cabeceras numéricas del panel de detalle sin re-renderizar todo el panel.
// IMPORTANTE: re-renderiza el contenido interno (no solo el text), para que la estructura
// editable/bloqueada cambie cuando se añade la primera línea o el primer recurso.
function _updateDetailHeader(nodo){
  const cod=nodo.codigo;
  const pc=curParent;
  const cantEl=document.getElementById('dh-cant');
  if(cantEl)cantEl.innerHTML=_renderHeaderCant(nodo,cod,pc);
  const precEl=document.getElementById('dh-precio');
  if(precEl)precEl.innerHTML=_renderHeaderPrecio(nodo,cod);
  const impEl=document.getElementById('dh-importe');
  if(impEl)impEl.textContent=nodo.importe_fmt+' €';
}

// Número europeo → float (para valores ya numéricos no usar parseNum)
function toNum(v){
  if(typeof v==='number')return v;
  return parseNum(String(v));
}

// ---- Tabulator — Descomposición ----
function _initTabDescomp(nodo,cod){
  if(_tabDescomp){try{_tabDescomp.destroy()}catch(e){}; _tabDescomp=null;}
  const el=document.getElementById('tab-descomp');
  if(!el||typeof Tabulator==='undefined')return;
  const TLBL={'1':'MO','2':'MQ','3':'MT','4':'AUX'};
  const TCLS={'1':'badge-mo','2':'badge-mq','3':'badge-mt','4':'badge-aux'};
  const data=(nodo.recursos||[]).map(r=>({...r}));

  // Guarda una fila nueva (_isNew) cuando el usuario rellena el codigo
  function _commitNewRecurso(row){
    const d=row.getData();
    if(!d.codigo||!d.codigo.trim())return;
    row.update({_isNew:false});
    api({accion:'add_recurso',codigo_partida:cod,codigo_recurso:d.codigo.trim(),
      rendimiento:parseNum(String(d.rendimiento||1))||1,
      precio:parseNum(String(d.precio||0))||0,
      unidad:d.unidad||'',resumen:d.resumen||''}).then(j=>{
      if(!j){row.update({_isNew:true});return;}
      calcSave(null,j);
    });
  }

  // Columna numérica con cellEdited PROPIO — sin mutatorEdit (más predecible)
  const numColDescomp=(title,field,width,onChange)=>({
    title,field,editor:'input',width,hozAlign:'right',cssClass:'tab-num',
    formatter:'money',formatterParams:{decimal:',',thousand:'.',symbol:'',precision:4},
    cellEdited:function(cell){
      const row=cell.getRow();const d=row.getData();
      // Convertir el string del editor a número
      const nVal=parseNum(String(cell.getValue()));
      const oVal=parseNum(String(cell.getOldValue()));
      // Reescribir la celda con el número (para que el formatter lo muestre bien)
      if(cell.getValue()!==nVal){
        // setValue(value, mutate=false): no dispara cellEdited otra vez
        try{cell.setValue(nVal,false);}catch(e){}
      }
      if(d._isNew){_commitNewRecurso(row);return;}
      if(nVal===oVal)return;
      onChange(d,nVal);
    }
  });

  _tabDescomp=new Tabulator(el,{
    data, index:'codigo', layout:'fitColumns',
    movableRows:true, selectableRows:true, reactiveData:false,
    editTriggerEvent:'dblclick',   // editar con doble clic; clic simple = seleccionar fila
    placeholder:'Sin recursos — pulsa Añadir',
    columnDefaults:{headerSort:false,resizable:false},
    columns:[
      {rowHandle:true,formatter:'handle',headerSort:false,width:26,minWidth:26},
      {title:'Código',field:'codigo',editor:'input',width:110,
       formatter:c=>`<span style="font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--text-dim)">${esc(c.getValue()||'')}</span>`,
       cellEdited:function(cell){
         const row=cell.getRow();const d=row.getData();
         const nv=cell.getValue();const ov=cell.getOldValue();
         if(d._isNew){_commitNewRecurso(row);return;}
         if(nv===ov||!nv||!nv.trim())return;
         silentSave({accion:'codigo',codigo_viejo:ov,codigo_nuevo:nv.trim()});
       }},
      {title:'Tipo',field:'tipo_fiebdc',editor:'list',width:70,
       editorParams:{values:{'1':'MO','2':'MQ','3':'MT','4':'AUX'},clearable:false},
       formatter:c=>{const v=c.getValue()||'3';return `<span class="tt-badge ${TCLS[v]||'badge-mt'}">${TLBL[v]||'MT'}</span>`;},
       cellEdited:function(cell){
         const d=cell.getRow().getData();
         if(d._isNew)return;
         if(cell.getValue()===cell.getOldValue())return;
         silentSave({accion:'tipo_recurso',codigo:d.codigo,tipo_fiebdc:cell.getValue()});
       }},
      {title:'Descripción',field:'resumen',editor:'input',
       cellEdited:function(cell){
         const d=cell.getRow().getData();
         if(d._isNew)return;
         if(cell.getValue()===cell.getOldValue())return;
         silentSave({accion:'resumen',codigo:d.codigo,valor:cell.getValue()});
       }},
      {title:'Ud',field:'unidad',editor:'input',width:58,
       cellEdited:function(cell){
         const d=cell.getRow().getData();
         if(d._isNew)return;
         if(cell.getValue()===cell.getOldValue())return;
         silentSave({accion:'unidad',codigo:d.codigo,valor:cell.getValue()});
       }},
      numColDescomp('Rendimiento','rendimiento',95,
        (d,nv)=>calcSave({accion:'rendimiento',codigo_padre:cod,codigo_hijo:d.codigo,valor:nv})),
      numColDescomp('Precio','precio',80,
        (d,nv)=>calcSave({accion:'precio',codigo:d.codigo,valor:nv})),
      {title:'Importe',field:'importe_fmt',editable:false,width:82,hozAlign:'right',cssClass:'tab-num',
       formatter:c=>`<span style="color:var(--green);font-weight:500">${esc(c.getValue()||'')}</span>`},
      {title:'',field:'_del',width:34,hozAlign:'center',headerSort:false,
       formatter:()=>'<button class="btn btn-sm btn-danger" style="padding:1px 5px">×</button>',
       cellClick:(e,cell)=>{
         const d=cell.getRow().getData();
         if(d._isNew){cell.getRow().delete();return;}
         if(!confirm('¿Quitar recurso '+d.codigo+'?'))return;
         api({accion:'eliminar_recurso',codigo_partida:cod,codigo_recurso:d.codigo}).then(j=>{
           if(!j)return;
           cell.getRow().delete();
           calcSave(null,j);
         });
       }},
    ],
    rowMoved:function(row){
      const rows=_tabDescomp.getRows();
      const idx=rows.indexOf(row);
      const antes_de=idx<rows.length-1?rows[idx+1].getData().codigo:null;
      const d=row.getData();
      if(!d._isNew)silentSave({accion:'reordenar_recurso',codigo_partida:cod,codigo_recurso:d.codigo,antes_de});
    },
  });
  _instalarSeleccionRango(_tabDescomp, el);
}

// Selección de filas estilo lista: clic = una sola, Shift = rango, Ctrl = alternar.
// Tabulator de serie solo ALTERNA fila a fila (sin rango), así que tomamos el
// control con un listener en fase de captura que bloquea su selección nativa.
function _instalarSeleccionRango(tab, el){
  let anchor=null;
  el.addEventListener('click',function(e){
    const rowEl=e.target.closest('.tabulator-row');
    if(!rowEl)return;
    // No interferir con el botón borrar (×) ni con el asa de arrastre
    if(e.target.closest('button')||e.target.closest('.tabulator-row-handle'))return;
    e.stopImmediatePropagation();   // impedir la selección nativa de Tabulator
    const rows=tab.getRows();
    const idx=rows.findIndex(r=>r.getElement()===rowEl);
    if(idx<0)return;
    if(e.shiftKey && anchor!=null && anchor<rows.length){
      const a=Math.min(idx,anchor), b=Math.max(idx,anchor);
      tab.deselectRow();
      for(let i=a;i<=b;i++) rows[i].select();
    }else if(e.ctrlKey||e.metaKey){
      rows[idx].toggleSelect(); anchor=idx;
    }else{
      tab.deselectRow(); rows[idx].select(); anchor=idx;
    }
  },true);
}

// ---- Tabulator — Mediciones ----
function _initTabMedic(nodo,cod,pc){
  if(_tabMedic){try{_tabMedic.destroy()}catch(e){}; _tabMedic=null;}
  const el=document.getElementById('tab-medic');
  if(!el||typeof Tabulator==='undefined')return;
  const data=(nodo.lineas_medicion||[]).map((ln,i)=>({...ln,_idx:i}));

  // Guarda una fila nueva (_isNew) en el servidor y actualiza Tabulator con el parcial calculado
  function _commitNewRow(row){
    row.update({_isNew:false});
    const rd=row.getData();
    api({accion:'add_linea_medicion',codigo_hijo:cod,codigo_padre:pc,
      comentario:rd.comentario||'',
      n_uds:rd.n_uds||0, longitud:rd.longitud||0,
      anchura:rd.anchura||0, altura:rd.altura||0,
    }).then(j=>{
      if(!j){row.update({_isNew:true});return;}
      const nNew=findNode(j.arbol,cod);
      if(nNew&&_tabMedic){
        const idx=(nNew.lineas_medicion||[]).length-1;
        if(idx>=0){const ln=nNew.lineas_medicion[idx];row.update({_idx:idx,subtotal:ln.subtotal,subtotal_fmt:ln.subtotal_fmt});}
      }
      calcSave(null,j);
    });
  }

  // mutatorEdit: cualquier string escrito por el usuario → número (europeo o anglosajón)
  const toNumMut=v=>parseNum(String(v));

  // Definición de columna numérica: money formatter + mutatorEdit + cellEdited
  const numCol=(title,field)=>({
    title,field,editor:'input',width:72,hozAlign:'right',cssClass:'tab-num',
    formatter:'money',
    formatterParams:{decimal:',',thousand:'.',symbol:'',precision:3},
    mutatorEdit:toNumMut,
    accessorClipboard:v=>(v==null?'':String(v).replace('.',',')),
    cellEdited:function(cell){
      const row=cell.getRow();const d=row.getData();
      const nv=cell.getValue();   // ya es número tras mutatorEdit
      const ov=cell.getOldValue();// también número
      if(d._isNew){_commitNewRow(row);return;}
      if(nv===ov)return;
      calcSave({accion:'medicion',codigo_hijo:cod,codigo_padre:pc,indice:d._idx,campo:field,valor:nv});
    }
  });

  // Parser TSV desde Excel
  function _parseMedClipboard(clipboard){
    const rows=clipboard.trim().split(/\r?\n/).filter(r=>r.trim());
    if(!rows.length)return[];
    const cols0=rows[0].split('\t');
    const start=(cols0.length>=2&&isNaN(parseFloat((cols0[1]||'').replace(',','.'))))?1:0;
    const baseIdx=_tabMedic?_tabMedic.getDataCount():data.length;
    return rows.slice(start).map((row,i)=>{
      const c=row.split('\t');
      return {_isNew:true,_idx:baseIdx+i,
        comentario:(c[0]||'').trim(),
        n_uds:c[1]?parseNum(c[1]):0,longitud:c[2]?parseNum(c[2]):0,
        anchura:c[3]?parseNum(c[3]):0,altura:c[4]?parseNum(c[4]):0,
        subtotal:0,subtotal_fmt:''};
    });
  }

  _tabMedic=new Tabulator(el,{
    data,index:'_idx',layout:'fitColumns',
    movableRows:true,selectableRows:true,reactiveData:false,
    editTriggerEvent:'dblclick',   // editar con doble clic; clic simple = seleccionar fila
    clipboard:'paste',clipboardCopyHeader:false,
    clipboardPasteAction:'insert',clipboardPasteParser:_parseMedClipboard,
    placeholder:'Sin mediciones — pulsa Añadir',
    columnDefaults:{headerSort:false,resizable:false},
    columns:[
      {rowHandle:true,formatter:'handle',headerSort:false,width:26,minWidth:26,clipboard:false},
      {title:'Comentario',field:'comentario',editor:'input',
       cellEdited:function(cell){
         const row=cell.getRow();const d=row.getData();
         const nv=cell.getValue();const ov=cell.getOldValue();
         if(d._isNew){_commitNewRow(row);return;}
         if(nv===ov)return;
         silentSave({accion:'medicion',codigo_hijo:cod,codigo_padre:pc,indice:d._idx,campo:'comentario',valor:nv});
       }},
      numCol('Uds','n_uds'),
      numCol('Largo','longitud'),
      numCol('Ancho','anchura'),
      numCol('Alto','altura'),
      {title:'Parcial',field:'subtotal_fmt',editable:false,width:82,hozAlign:'right',cssClass:'tab-num',clipboard:false,
       formatter:c=>`<strong style="color:var(--green)">${esc(c.getValue()||'')}</strong>`},
      {title:'',field:'_del',width:34,hozAlign:'center',clipboard:false,
       formatter:()=>'<button class="btn btn-sm btn-danger" style="padding:1px 5px">×</button>',
       cellClick:(e,cell)=>{
         const d=cell.getRow().getData();
         if(d._isNew){cell.getRow().delete();return;}
         if(!confirm('¿Eliminar línea?'))return;
         api({accion:'eliminar_linea_medicion',codigo_hijo:cod,codigo_padre:pc,indice:d._idx}).then(j=>{
           if(!j)return;
           cell.getRow().delete();
           _tabMedic.getRows().forEach((r,i)=>r.update({_idx:i}));
           calcSave(null,j);
         });
       }},
    ],
    rowMoved:function(row){
      const from=row.getData()._idx;
      const to=_tabMedic.getRows().indexOf(row);
      if(from!==to)api({accion:'reordenar_medicion',codigo_hijo:cod,codigo_padre:pc,from_idx:from,to_idx:to}).then(j=>{
        if(!j)return;
        _tabMedic.getRows().forEach((r,i)=>r.update({_idx:i}));
        calcSave(null,j);
      });
    },
  });

  // Guardar filas pegadas desde Excel y recargar con subtotales del servidor
  _tabMedic.on('clipboardPasted',async function(_clip,rowData){
    const nuevas=rowData.filter(r=>r._isNew);
    if(!nuevas.length)return;
    let lastJ=null;
    for(const r of nuevas){
      const j=await api({accion:'add_linea_medicion',codigo_hijo:cod,codigo_padre:pc,
        comentario:r.comentario||'',n_uds:r.n_uds||0,longitud:r.longitud||0,
        anchura:r.anchura||0,altura:r.altura||0});
      if(j)lastJ=j;
    }
    if(lastJ){
      const nNew=findNode(lastJ.arbol,cod);
      if(nNew&&_tabMedic)_tabMedic.replaceData((nNew.lineas_medicion||[]).map((ln,i)=>({...ln,_idx:i})));
      calcSave(null,lastJ);
    }
  });
  _instalarSeleccionRango(_tabMedic, el);
}

// ---- Render app ----
function renderApp(){
  document.getElementById('uploadScreen').style.display='none';
  document.getElementById('mainApp').style.display='';
  document.getElementById('headerActions').style.display='';
  renderStats();renderTree();
  document.getElementById('detailPanel').innerHTML='<div class="detail-empty">Selecciona una partida</div>';
}
// Estado de plegado del árbol: codigos abiertos
let openCaps=new Set();

// ---- Drag & drop state ----
let _dragInfo=null;  // {codigo, padre}
let _dropInfo=null;  // {padre_destino, antes_de}   antes_de=null → append

function _cleanDrop(){
  document.querySelectorAll('.drop-before,.drop-after,.drop-into')
    .forEach(el=>el.classList.remove('drop-before','drop-after','drop-into'));
  _dropInfo=null;
}

// Devuelve el código del hermano SIGUIENTE de targetCod dentro de parentCod,
// o null si targetCod es el último hijo.
function _siblingAfter(parentCod, targetCod){
  let arr;
  if(!parentCod){arr=treeData}
  else{const n=findNode(treeData,parentCod);arr=n?n.hijos:[]}
  const i=arr.findIndex(n=>n.codigo===targetCod);
  return(i>=0&&i<arr.length-1)?arr[i+1].codigo:null;
}
function renderStats(){
  const i=fileInfo;
  document.getElementById('statsBar').innerHTML=`
    <div class="stat"><span class="stat-label">Obra</span><span class="stat-value">${esc(i.obra)}</span></div>
    <div class="stat"><span class="stat-label">Versión</span><span class="stat-value">${esc(i.version)}</span></div>
    <div class="stat"><span class="stat-label">Programa</span><span class="stat-value">${esc(i.programa)}</span></div>
    <div class="stat"><span class="stat-label">Capítulos</span><span class="stat-value">${i.capitulos}</span></div>
    <div class="stat"><span class="stat-label">Partidas</span><span class="stat-value">${i.partidas}</span></div>
    <div class="stat"><span class="stat-label">PEM</span><span class="stat-value total">${esc(i.total_fmt)} €</span></div>`;
  // Banner si el archivo está en carpeta temporal
  const tb=document.getElementById('tempBanner');
  if(i.archivo_temporal){
    tb.style.display='';
  }else{
    tb.style.display='none';
  }
}
function renderTree(){
  const tb=document.getElementById('treeBody');tb.innerHTML='';
  treeData.forEach(n=>appendTreeRows(tb,n,0,''));
  // Fila fantasma a nivel raíz para añadir capítulos de primer nivel
  tb.appendChild(mkGhostTreeRow('',0));
}
// Enfoca y activa la celda de código de la ghost row correspondiente a codPadre
function focusGhostRow(codPadre){
  const ghost=[...document.querySelectorAll('.ghost-row')]
    .find(g=>g.dataset.ghostParent===(codPadre||''));
  if(!ghost)return;
  const cell=ghost.querySelector('.ecell[data-editable=true]');
  if(!cell)return;
  ghost.scrollIntoView({behavior:'smooth',block:'center'});
  // setTimeout para evitar que el ecActivate se interrumpa con el scroll
  setTimeout(()=>ecActivate(cell),120);
}
// Inserta una fila por nodo (capítulo o partida) y, si está abierto, sus hijos.
function appendTreeRows(tbody,nodo,level,parentCod){
  tbody.appendChild(mkTreeRow(nodo,level,parentCod));
  const isCap=nodo.tipo==='capitulo';
  if(isCap && openCaps.has(nodo.codigo)){
    if(nodo.hijos && nodo.hijos.length>0)
      nodo.hijos.forEach(h=>appendTreeRows(tbody,h,level+1,nodo.codigo));
    tbody.appendChild(mkGhostTreeRow(nodo.codigo,level+1));
  }
}
// Fila fantasma al final de cada capítulo: fila en blanco idéntica a una fila real.
// Tab navega entre celdas SIN crear el concepto.
// Intro (Enter) en cualquier celda → crea el concepto con todo lo rellenado.
// Código que termina en # → capítulo; sin # → partida.
function mkGhostTreeRow(codigoPadre,level){
  const tr=document.createElement('tr');
  tr.className='ghost-row';
  tr.dataset.ghostParent=codigoPadre||'';
  tr.dataset.rowKey=`t-ghost:${codigoPadre||''}`;
  const pad=level*16;
  const indentHtml=`<span class="tt-indent" style="width:${pad}px"></span>`;

  // Borrador acumulado hasta que el usuario confirma
  const draft={codigo:'',resumen:'',unidad:'',precio:0};
  const codHtml=ec('',true,v=>{draft.codigo=(v||'').trim();},false);
  const resHtml=ec('',true,v=>{draft.resumen=(v||'').trim();},false);
  const udHtml =ec('',true,v=>{draft.unidad =(v||'').trim();},false);
  const preHtml=ec('',true,v=>{draft.precio =parseNum(v);  },true);

  tr.innerHTML=
    `<td class="col-cod cod">${codHtml}</td>`+
    `<td class="col-resumen">${indentHtml}`+
      `<span class="tt-toggle leaf">▶</span>`+
      `<span class="tt-badge" style="visibility:hidden">Cap</span>`+
      `${resHtml}</td>`+
    `<td class="col-ud">${udHtml}</td>`+
    `<td class="num col-cant"><span class="ecell num"></span></td>`+
    `<td class="num col-precio">${preHtml}</td>`+
    `<td class="num col-imp"><span class="ecell num"></span></td>`+
    `<td class="col-act" style="color:var(--text-muted);font-size:10px;padding-right:6px;text-align:right">↵ Intro</td>`;

  _wireGhostRow(tr,()=>{
    const cod=draft.codigo.trim();
    if(!cod)return;
    openCaps.add(codigoPadre);
    api(cod.endsWith('#')
      ?{accion:'add_capitulo',codigo:cod,resumen:draft.resumen,codigo_padre:codigoPadre}
      :{accion:'add_partida',codigo_padre:codigoPadre,codigo:cod,
        unidad:draft.unidad,resumen:draft.resumen,precio:draft.precio}
    ).then(refresh);
  });

  // La fila fantasma también es zona de drop: soltar aquí → append al capítulo padre
  tr.addEventListener('dragover',e=>{
    if(!_dragInfo)return;
    e.preventDefault();e.dataTransfer.dropEffect='move';
    _cleanDrop();tr.classList.add('drop-into');
    _dropInfo={padre_destino:codigoPadre,antes_de:null};
  });
  tr.addEventListener('dragleave',e=>{if(!tr.contains(e.relatedTarget))tr.classList.remove('drop-into')});
  tr.addEventListener('drop',e=>{
    e.preventDefault();
    if(!_dragInfo||!_dropInfo)return;
    const{codigo,padre:padre_origen}=_dragInfo;
    _cleanDrop();_dragInfo=null;
    api({accion:'mover',codigo,padre_origen,padre_destino:codigoPadre,antes_de:null}).then(refresh);
  });
  return tr;
}
// Crea una fila <tr> con sus celdas y handlers.
function mkTreeRow(nodo,level,parentCod){
  const tr=document.createElement('tr');
  const isCap=nodo.tipo==='capitulo';
  const hasKids=isCap && nodo.hijos && nodo.hijos.length>0;
  const isOpen=openCaps.has(nodo.codigo);
  tr.className='row-'+(isCap?'cap':'part');
  if(curNode && curNode.codigo===nodo.codigo && curParent===parentCod) tr.classList.add('active');
  tr._nodoData=nodo;tr._parentCod=parentCod;
  tr.dataset.rowKey=`t:${parentCod}>${nodo.codigo}`;

  // Sangría (16px por nivel) + toggle ▶
  const pad=level*16;
  const toggleHtml=`<span class="tt-toggle ${hasKids?'':'leaf'}${isOpen?' open':''}" data-act="toggle">▶</span>`;
  const indentHtml=`<span class="tt-indent" style="width:${pad}px"></span>`;
  const TIPOS_ARBOL=[{value:'capitulo',label:'Cap'},{value:'partida',label:'Part'}];
  const badgeHtml=ecSelect(
    isCap?'capitulo':'partida',
    TIPOS_ARBOL,
    v=>api({accion:'cambiar_tipo',codigo:nodo.codigo,tipo:v}).then(refresh),
    isCap?'badge-cap':'badge-part'
  );

  // Cantidad: capítulo → vacío; partida con líneas → total calculado; partida sin líneas → editable
  const numLineas=nodo.lineas_medicion?nodo.lineas_medicion.length:0;
  let cantHtml;
  if(isCap){
    cantHtml=``;   // los capítulos no tienen cantidad propia
  }else if(numLineas>0){
    cantHtml=`<span title="Suma de líneas de medición">${esc(nodo.medicion_fmt)}</span>`;
  }else{
    cantHtml=ec(nodo.medicion_fmt||'0',true,v=>setCantidadSimple(nodo.codigo,parentCod,parseNum(v)),true);
  }

  // Acciones contextuales por fila
  let actsHtml='<div class="row-actions">';
  if(parentCod) actsHtml+=`<span class="drag-handle" title="Arrastrar para mover">⠿</span>`;
  if(isCap){
    actsHtml+=`<button title="+ Subcapítulo" data-act="addcap">＋</button>`;
    actsHtml+=`<button title="+ Partida" data-act="addpart">▦</button>`;
  }
  if(parentCod) actsHtml+=`<button class="danger" title="Eliminar" data-act="del">×</button>`;
  actsHtml+='</div>';

  // El atributo draggable solo si tiene padre (no se puede mover la raíz)
  if(parentCod) tr.draggable=true;

  // Precio unitario: capítulo → vacío; partida con desglose → bloqueado; partida simple → editable
  let precioHtml;
  if(isCap){
    precioHtml=`<span class="ecell num" style="color:var(--text-muted)">—</span>`;
  }else if(nodo.recursos&&nodo.recursos.length>0){
    precioHtml=`<span class="ecell num" contenteditable="false"
      title="Calculado desde la descomposición — no editable directamente"
      style="cursor:default;opacity:.7">${esc(nodo.precio_fmt)}</span>`;
  }else{
    precioHtml=ec(nodo.precio_fmt||'0,00',true,v=>api({accion:'precio',codigo:nodo.codigo,valor:parseNum(v)}).then(refresh),true);
  }

  tr.innerHTML=
    `<td class="col-cod cod">${ec(nodo.codigo,true,v=>api({accion:'codigo',codigo_viejo:nodo.codigo,codigo_nuevo:v.trim()}).then(refresh),false)}</td>`+
    `<td class="col-resumen" title="${esc(nodo.resumen||'')}">${indentHtml}${toggleHtml}${badgeHtml}${ec(nodo.resumen,true,v=>api({accion:'resumen',codigo:nodo.codigo,valor:v}).then(refresh),false)}</td>`+
    `<td class="col-ud">${isCap?'':ec(nodo.unidad||'',true,v=>api({accion:'unidad',codigo:nodo.codigo,valor:v}).then(refresh),false)}</td>`+
    `<td class="num col-cant">${cantHtml}</td>`+
    `<td class="num col-precio">${precioHtml}</td>`+
    `<td class="num col-imp imp">${esc(nodo.importe_fmt)} €</td>`+
    `<td class="col-act">${actsHtml}</td>`;

  // ---- Clic ----
  tr.addEventListener('click',e=>{
    const t=e.target;
    const act=t.getAttribute && t.getAttribute('data-act');
    if(act==='toggle'){ toggleCap(nodo.codigo); return; }
    if(act==='addcap'){ e.stopPropagation(); addCapitulo(nodo.codigo); return; }
    if(act==='addpart'){ e.stopPropagation(); addPartida(nodo.codigo); return; }
    if(act==='del'){ e.stopPropagation(); eliminarConcepto(nodo.codigo,parentCod); return; }

    // ---- Selección múltiple ----
    if(e.ctrlKey||e.metaKey){
      // Ctrl+clic → toggle en la selección
      e.preventDefault();
      const key=`${parentCod}>${nodo.codigo}`;
      const idx=selectedNodes.findIndex(s=>s.codigo===nodo.codigo&&s.padre===parentCod);
      if(idx>=0){ selectedNodes.splice(idx,1); tr.classList.remove('tree-selected'); }
      else{ selectedNodes.push({codigo:nodo.codigo,padre:parentCod}); tr.classList.add('tree-selected'); }
      return;
    }
    if(e.shiftKey&&_selectAnchor){
      // Shift+clic → rango desde el ancla hasta aquí
      e.preventDefault();
      const allRows=[...document.querySelectorAll('.ttable tbody tr[data-row-key^="t:"]')];
      const anchorKey=`t:${_selectAnchor.padre}>${_selectAnchor.codigo}`;
      const curKey=`t:${parentCod}>${nodo.codigo}`;
      const ai=allRows.findIndex(r=>r.dataset.rowKey===anchorKey);
      const ci=allRows.findIndex(r=>r.dataset.rowKey===curKey);
      if(ai>=0&&ci>=0){
        const [from,to]=[Math.min(ai,ci),Math.max(ai,ci)];
        selectedNodes=[];
        document.querySelectorAll('.ttable tr.tree-selected').forEach(r=>r.classList.remove('tree-selected'));
        allRows.slice(from,to+1).forEach(r=>{
          r.classList.add('tree-selected');
          const rk=r.dataset.rowKey.replace(/^t:/,'');
          const sep=rk.indexOf('>');
          selectedNodes.push({padre:rk.slice(0,sep),codigo:rk.slice(sep+1)});
        });
      }
      return;
    }

    // Clic normal → limpia selección múltiple
    selectedNodes=[];
    document.querySelectorAll('.ttable tr.tree-selected').forEach(r=>r.classList.remove('tree-selected'));
    _selectAnchor={codigo:nodo.codigo,padre:parentCod};

    const sameNode=curNode&&curNode.codigo===nodo.codigo&&curParent===parentCod;
    document.querySelectorAll('.ttable tbody tr.active').forEach(x=>x.classList.remove('active'));
    tr.classList.add('active');
    curNode=nodo; curParent=parentCod;
    if(sameNode) return;
    if(isCap){
      renderDetail(nodo);
      if(hasKids) toggleCap(nodo.codigo);
    }else{
      renderDetail(nodo);
    }
  });

  // ---- Drag ----
  if(parentCod){
    tr.addEventListener('dragstart',e=>{
      if(document.activeElement&&document.activeElement.contentEditable==='true'){e.preventDefault();return}
      _dragInfo={codigo:nodo.codigo,padre:parentCod};
      e.dataTransfer.effectAllowed='move';
      e.dataTransfer.setData('text/plain',nodo.codigo);
      // Aplazar el addClass para que el snapshot del drag muestre la fila normal
      requestAnimationFrame(()=>tr.classList.add('dragging'));
    });
    tr.addEventListener('dragend',()=>{tr.classList.remove('dragging');_cleanDrop();_dragInfo=null});
  }

  tr.addEventListener('dragover',e=>{
    if(!_dragInfo||_dragInfo.codigo===nodo.codigo) return;
    e.preventDefault();e.dataTransfer.dropEffect='move';
    _cleanDrop();
    const mid=tr.getBoundingClientRect().top+tr.getBoundingClientRect().height/2;
    if(isCap&&e.clientY>mid){
      // Hover sobre la mitad inferior de un capítulo → soltar DENTRO
      tr.classList.add('drop-into');
      _dropInfo={padre_destino:nodo.codigo,antes_de:null};
    }else if(e.clientY<mid){
      tr.classList.add('drop-before');
      _dropInfo={padre_destino:parentCod,antes_de:nodo.codigo};
    }else{
      tr.classList.add('drop-after');
      const sig=_siblingAfter(parentCod,nodo.codigo);
      _dropInfo={padre_destino:parentCod,antes_de:sig};
    }
  });
  tr.addEventListener('dragleave',e=>{
    if(!tr.contains(e.relatedTarget))tr.classList.remove('drop-before','drop-after','drop-into');
  });
  tr.addEventListener('drop',e=>{
    e.preventDefault();
    if(!_dragInfo||!_dropInfo)return;
    const {padre_destino,antes_de}=_dropInfo;
    const {codigo,padre:padre_origen}=_dragInfo;
    _cleanDrop();_dragInfo=null;
    api({accion:'mover',codigo,padre_origen,padre_destino,antes_de}).then(refresh);
  });

  return tr;
}
function toggleCap(codigo){
  if(openCaps.has(codigo)) openCaps.delete(codigo); else openCaps.add(codigo);
  renderTree();
}
// Cuando la partida no tiene líneas, editar la cantidad crea una línea simple.
async function setCantidadSimple(codHijo,codPadre,valor){
  if(!codPadre) return;
  refresh(await api({accion:'add_linea_medicion',codigo_hijo:codHijo,codigo_padre:codPadre,
    comentario:'',n_uds:valor,longitud:0,anchura:0,altura:0}));
}

// ---- Editable cell helper ----
// Las celdas son read-only por defecto. Doble clic activa la edición.
function ec(val,editable,saveCallback,isNum){
  const cls=isNum?'ecell num':'ecell';
  const display=val==null||val===''?'':val;
  if(!editable||!saveCallback){
    return `<span class="${cls}" contenteditable="false">${esc(String(display))}</span>`;
  }
  const id='ec'+Math.random().toString(36).substr(2,6);
  window['_ec_'+id]=saveCallback;
  const orig=esc(String(display));
  return `<span class="${cls}" contenteditable="false" data-editable="true" data-orig="${orig}" id="${id}"
    title="Doble clic para editar"
    ondblclick="ecActivate(this)"
    onblur="ecBlur(this,'${id}')"
    onkeydown="ecKey(event,this)"
    tabindex="0">${esc(String(display))}</span>`;
}

// Activa la edición de una celda (doble clic o Tab+Enter)
function ecActivate(el){
  if(el.dataset.editable!=='true')return;
  el.dataset.orig=el.textContent;   // guarda valor actual para Escape
  el.contentEditable='true';
  el.title='';
  el.focus();
  const range=document.createRange();
  range.selectNodeContents(el);
  const sel=window.getSelection();
  sel.removeAllRanges();sel.addRange(range);
}

// Desactiva la edición y guarda solo si el valor cambió
function ecBlur(el,id){
  if(el.dataset.cancelling)return;
  el.contentEditable='false';
  el.title='Doble clic para editar';
  if(el.textContent===el.dataset.orig)return;   // sin cambios → no disparar API
  if(window['_ec_'+id])window['_ec_'+id](el.textContent);
}

// Conecta una "ghost row" (fila para añadir un nuevo registro) con sus comportamientos:
//   - Tab/Shift+Tab navegan entre las celdas de la fila sin crear nada.
//   - Tab desde la ÚLTIMA celda → commit.
//   - Enter en cualquier celda → commit.
//   - Foco fuera de toda la fila → commit.
//   - Escape → cancela la edición de la celda.
// commitFn se invoca sin parámetros; debe leer su propio borrador y enviar la petición.
function _wireGhostRow(tr,commitFn){
  let _done=false;
  const fire=()=>{if(_done)return;_done=true;commitFn();};
  const cells=[...tr.querySelectorAll('.ecell[data-editable=true]')];
  cells.forEach((el,idx)=>{
    el.onkeydown=function(e){
      if(e.key==='Enter'){
        e.preventDefault();
        if(el.contentEditable==='true')el.blur();   // dispara ecBlur → guarda en borrador
        fire();
        return;
      }
      if(el.contentEditable!=='true')return;
      if(e.key==='Tab'){
        e.preventDefault();el.blur();
        const tgt=e.shiftKey?cells[idx-1]:cells[idx+1];
        if(tgt){ecActivate(tgt);}
        else if(!e.shiftKey){fire();}   // Tab desde la última celda → commit
        return;
      }
      if(e.key==='ArrowUp'||e.key==='ArrowDown'){
        // Navega a la misma columna en la fila anterior/siguiente de la misma tabla
        const _tr=el.closest('tr');
        const _tbl=el.closest('table');
        if(!_tr||!_tbl)return;
        const rows=[..._tbl.querySelectorAll('tr')].filter(r=>r.querySelector('.ecell[data-editable=true]'));
        const ri=rows.indexOf(_tr);
        const rc=[..._tr.querySelectorAll('.ecell[data-editable=true]')];
        const ci=rc.indexOf(el);
        const tgtRow=e.key==='ArrowDown'?rows[ri+1]:rows[ri-1];
        if(!tgtRow)return;
        const tc=[...tgtRow.querySelectorAll('.ecell[data-editable=true]')];
        const tgt=tc[ci]||tc[tc.length-1];
        if(tgt){e.preventDefault();el.blur();ecActivate(tgt);}
        return;
      }
      if(e.key==='Escape'){
        e.preventDefault();
        el.dataset.cancelling='true';
        el.textContent=el.dataset.orig||'';
        el.contentEditable='false';
        el.title='Doble clic para editar';
        el.blur();
        delete el.dataset.cancelling;
      }
    };
  });
  tr.addEventListener('focusout',()=>{
    setTimeout(()=>{
      if(!tr.isConnected||tr.contains(document.activeElement))return;
      fire();
    },50);
  });
}

function ecKey(e,el){
  if(el.contentEditable!=='true')return;

  if(e.key==='Escape'){
    e.preventDefault();
    el.dataset.cancelling='true';
    el.textContent=el.dataset.orig||'';
    el.contentEditable='false';
    el.title='Doble clic para editar';
    el.blur();
    delete el.dataset.cancelling;
    return;
  }

  // Helpers: celdas editables de la fila actual y de la tabla actual
  const _tr=el.closest('tr');
  const _table=el.closest('table');
  const _rowCells=()=>_tr?[..._tr.querySelectorAll('.ecell[data-editable=true]')]:[];
  const _tableRows=()=>_table?[..._table.querySelectorAll('tr')].filter(r=>r.querySelector('.ecell[data-editable=true]')):[];

  // Enter → mueve a la celda de abajo (estilo Excel)
  if(e.key==='Enter'){
    e.preventDefault();
    const rc=_rowCells();const ci=rc.indexOf(el);
    const rows=_tableRows();const ri=rows.indexOf(_tr);
    const tgtRow=rows[ri+1];
    el.blur();   // guarda valor → puede disparar refresh; el foco se restaura por _restoreFocusPos
    if(tgtRow){
      const tc=[...tgtRow.querySelectorAll('.ecell[data-editable=true]')];
      const tgt=tc[ci]||tc[tc.length-1];
      if(tgt)ecActivate(tgt);
    }
    return;
  }

  if(e.key==='Tab'){
    e.preventDefault();el.blur();
    const rc=_rowCells();const ci=rc.indexOf(el);
    if(!e.shiftKey){
      // → siguiente celda en la misma fila; si es la última, primera celda de la fila siguiente
      if(ci<rc.length-1){ecActivate(rc[ci+1]);}
      else{
        const rows=_tableRows();const ri=rows.indexOf(_tr);
        const nxt=rows[ri+1];
        if(nxt){const nc=[...nxt.querySelectorAll('.ecell[data-editable=true]')];if(nc.length)ecActivate(nc[0]);}
      }
    }else{
      // ← celda anterior en la misma fila; si es la primera, última celda de la fila anterior
      if(ci>0){ecActivate(rc[ci-1]);}
      else{
        const rows=_tableRows();const ri=rows.indexOf(_tr);
        const prv=rows[ri-1];
        if(prv){const pc=[...prv.querySelectorAll('.ecell[data-editable=true]')];if(pc.length)ecActivate(pc[pc.length-1]);}
      }
    }
    return;
  }

  if(e.key==='ArrowDown'||e.key==='ArrowUp'){
    const rc=_rowCells();const ci=rc.indexOf(el);
    const rows=_tableRows();const ri=rows.indexOf(_tr);
    const tgtRow=e.key==='ArrowDown'?rows[ri+1]:rows[ri-1];
    if(!tgtRow)return;
    const tc=[...tgtRow.querySelectorAll('.ecell[data-editable=true]')];
    const tgtCell=tc[ci]||tc[tc.length-1];
    if(tgtCell){e.preventDefault();el.blur();ecActivate(tgtCell);}
    return;
  }

  if(e.key==='ArrowLeft'||e.key==='ArrowRight'){
    // Navega a la celda adyacente de la misma fila solo al llegar al borde del texto
    const sel=window.getSelection();
    const atStart=sel&&sel.anchorOffset===0;
    const atEnd=sel&&sel.anchorOffset===el.textContent.length;
    if(e.key==='ArrowLeft'&&!atStart)return;
    if(e.key==='ArrowRight'&&!atEnd)return;
    e.preventDefault();el.blur();
    const rc=_rowCells();const ci=rc.indexOf(el);
    const next=e.key==='ArrowRight'?rc[ci+1]:rc[ci-1];
    if(next)ecActivate(next);
    return;
  }
}

// ---- ecSelect: badge con desplegable de opciones (doble clic) ----
// options: [{value, label}]   extraClass: clase CSS extra del badge (badge-cap, badge-mo, ...)
function ecSelect(val, options, saveCallback, extraClass){
  const id='ecs'+Math.random().toString(36).substr(2,6);
  window['_ecs_'+id]={cb:saveCallback, opts:options};
  const cur=options.find(o=>o.value===val)||options[0];
  return `<span class="tt-badge ${extraClass||''} clickable" id="${id}"
    title="Doble clic para cambiar tipo" data-val="${esc(String(val))}" data-extra="${esc(extraClass||'')}"
    ondblclick="event.stopPropagation();ecSelectActivate('${id}')">${esc(cur.label)}</span>`;
}

function ecSelectActivate(id){
  const el=document.getElementById(id);
  if(!el||!window['_ecs_'+id])return;
  // Cerrar cualquier otro desplegable abierto
  document.querySelectorAll('.tipo-dropdown').forEach(d=>d.remove());
  const{cb,opts}=window['_ecs_'+id];
  const currentVal=el.dataset.val;
  const dd=document.createElement('div');
  dd.className='tipo-dropdown';
  const rect=el.getBoundingClientRect();
  dd.style.cssText=`position:fixed;top:${rect.bottom+2}px;left:${rect.left}px`;
  // Mapa de label → clase badge para actualizar el color al seleccionar
  const BADGE_CLS={Cap:'badge-cap',Part:'badge-part',MO:'badge-mo',MQ:'badge-mq',MT:'badge-mt',AUX:'badge-aux'};
  opts.forEach(o=>{
    const item=document.createElement('div');
    item.className='tipo-dropdown-item'+(o.value===currentVal?' active':'');
    item.textContent=o.label;
    item.addEventListener('click',e=>{
      e.stopPropagation();
      dd.remove();
      el.textContent=o.label;
      el.dataset.val=o.value;
      // Actualizar clase de color del badge
      const newCls=BADGE_CLS[o.label]||el.dataset.extra||'';
      el.className=el.className.replace(/badge-\w+/g,'').replace(/\s+/g,' ').trim();
      if(newCls)el.classList.add(newCls);
      el.dataset.extra=newCls;
      cb(o.value);
    });
    dd.appendChild(item);
  });
  document.body.appendChild(dd);
  setTimeout(()=>{
    document.addEventListener('click',function h(e){
      if(!dd.contains(e.target)){dd.remove();document.removeEventListener('click',h)}
    });
  },0);
}

// ---- Parse number from cell ----
// Soporta formato europeo (1.234,56) y anglosajón/JS (1234.56 ó 1.5).
// Heurística: si hay coma → coma=decimal, puntos=miles (europeo).
//             si no hay coma → punto=decimal (JS/anglosajón).
function parseNum(s){
  const str=String(s==null?'':s).trim();
  if(!str||str==='-')return 0;
  if(str.includes(','))
    // Europeo: quita puntos-miles, convierte coma-decimal a punto
    return parseFloat(str.replace(/\./g,'').replace(',','.'))||0;
  // Sin coma: el punto es separador decimal (formato JS / BC3 anglosajón)
  return parseFloat(str)||0;
}

// ---- Render detail (partidas) ----
function renderDetail(nodo){
  const panel=document.getElementById('detailPanel');
  if(nodo.tipo==='capitulo'){
    if(_pasteHandler){document.removeEventListener('paste',_pasteHandler);_pasteHandler=null}
    panel.innerHTML='<div class="detail-empty">Selecciona una partida</div>';
    return;
  }
  const pc=curParent;
  const cod=nodo.codigo;

  let h=`<div class="detail-header">
    <div class="detail-code">${ec(cod,true,v=>api({accion:'codigo',codigo_viejo:cod,codigo_nuevo:v.trim()}).then(refresh),false)}</div>
    <div class="detail-name">${ec(nodo.resumen,true,v=>api({accion:'resumen',codigo:cod,valor:v}).then(refresh),false)}</div>
    <div class="detail-meta">
      <div class="detail-meta-item"><span class="detail-meta-label">Unidad</span><span class="detail-meta-value">${ec(nodo.unidad,true,v=>api({accion:'unidad',codigo:cod,valor:v}).then(refresh),false)}</span></div>
      <div class="detail-meta-item"><span class="detail-meta-label">Cantidad</span>
        <span class="detail-meta-value" id="dh-cant">${_renderHeaderCant(nodo,cod,pc)}</span>
      </div>
      <div class="detail-meta-item"><span class="detail-meta-label">Precio</span>
        <span class="detail-meta-value" id="dh-precio">${_renderHeaderPrecio(nodo,cod)}</span>
      </div>
      <div class="detail-meta-item"><span class="detail-meta-label">Importe</span><span class="detail-meta-value green" id="dh-importe">${esc(nodo.importe_fmt)} €</span></div>
    </div></div>`;

  // Texto (descripción larga)
  h+=`<div class="detail-section"><h3>Descripción</h3>
    <div class="detail-text" id="detailTexto" contenteditable="true"
      data-placeholder="Haz clic para añadir una descripción detallada…"
      data-orig="${esc(nodo.texto||'')}"
    >${esc(nodo.texto||'')}</div></div>`;

  // Descomposición — Tabulator
  h+=`<div class="detail-section"><h3>Descomposición
      <span style="font-size:10px;color:var(--text-muted);font-weight:400;margin-left:6px">Clic = seleccionar (Shift = rango) · Doble clic = editar · Ctrl+C copia · Supr borra</span>
    </h3>
    <div id="tab-descomp"></div>
    <div class="tab-actions">
      <span class="tab-add-row" onclick="
        if(!_tabDescomp)return;
        _tabDescomp.addRow({tipo_fiebdc:'3',_isNew:true},false).then(r=>{
          const ed=r.getCells().find(c=>c.getField()==='codigo');
          if(ed)setTimeout(()=>ed.edit(),30);
        })">＋ Añadir recurso</span>
      <span class="tab-del-row" onclick="eliminarSelDesglose()">🗑 Eliminar seleccionadas</span>
    </div>
  </div>`;

  // Mediciones — Tabulator
  if(pc){
    const medId='med-'+cod.replace(/[^a-zA-Z0-9]/g,'_');
    h+=`<div class="detail-section"><h3>Mediciones
      <span style="font-size:10px;color:var(--text-muted);font-weight:400;margin-left:6px">Clic = seleccionar (Shift = rango) · Doble clic = editar · Ctrl+V pega desde Excel · Supr borra</span>
    </h3>
    <div id="tab-medic" data-medid="${esc(medId)}"></div>
    <div class="tab-actions">
      <span class="tab-add-row" onclick="
        if(!_tabMedic)return;
        _tabMedic.addRow({_isNew:true,_idx:_tabMedic.getDataCount()},false).then(r=>{
          const ed=r.getCells().find(c=>c.getField()==='comentario');
          if(ed)setTimeout(()=>ed.edit(),30);
        })">＋ Añadir línea</span>
      <span class="tab-del-row" onclick="eliminarSelMedic()">🗑 Eliminar seleccionadas</span>
    </div>
    </div>`;
  }
  // Destruir instancias anteriores antes de reemplazar el DOM
  _destroyTabs();
  panel.innerHTML=h;

  // Descripción larga editable
  const textoEl=panel.querySelector('#detailTexto');
  if(textoEl){
    const _orig=nodo.texto||'';
    textoEl.addEventListener('blur',()=>{
      const val=textoEl.innerText.replace(/\r/g,'').trimEnd();
      if(val===_orig)return;
      api({accion:'texto',codigo:cod,valor:val});
    });
    textoEl.addEventListener('keydown',e=>{
      if(e.key==='Escape'){e.preventDefault();textoEl.innerText=_orig;textoEl.blur();}
    });
  }

  // Inicializar Tabulator (después de que el DOM exista)
  _initTabDescomp(nodo,cod);
  if(pc) _initTabMedic(nodo,cod,pc);
  // Limpiar cualquier handler de paste anterior (Tabulator gestiona el clipboard ahora)
  if(_pasteHandler){document.removeEventListener('paste',_pasteHandler);_pasteHandler=null;}
}

// Handler de pegado desde Excel — se guarda para poder desinstalarlo al cambiar de partida.
let _pasteHandler=null;
function _instalarPasteHandler(cod, pc){
  // Desinstalar el anterior si existe
  if(_pasteHandler){document.removeEventListener('paste',_pasteHandler);_pasteHandler=null}
  if(!pc) return;   // sin padre (raíz) no tiene mediciones
  _pasteHandler=function(e){
    // Sólo actuar si NO estamos editando un ecell (para no interferir con el pegado normal en celdas)
    if(document.activeElement&&document.activeElement.contentEditable==='true') return;
    const panel=document.getElementById('detailPanel');
    if(!panel||!panel.querySelector('#med-'+cod.replace(/[^a-zA-Z0-9]/g,'_'))) return; // partida ya no está visible
    const txt=(e.clipboardData||window.clipboardData).getData('text');
    if(!txt) return;
    // Detectar si es pegado TSV de Excel (con tabuladores o al menos una línea de datos)
    const rows=txt.trim().split(/\r?\n/).filter(r=>r.trim());
    if(!rows.length) return;
    // Si no hay tabuladores y sólo hay una fila podría ser texto normal — sólo interceptamos si hay tabs o varias filas
    if(rows.length===1 && !rows[0].includes('\t')) return;
    e.preventDefault();
    const cols0=rows[0].split('\t');
    // Heurística: si la primera columna parece una cabecera (todo texto, sin números) la saltamos
    const primeraEsHeader=cols0.length>=2 && isNaN(parseFloat(cols0[1].replace(',','.')));
    const filas=primeraEsHeader?rows.slice(1):rows;
    if(!filas.length) return;
    let chain=Promise.resolve(), lastJ=null;
    filas.forEach(row=>{
      const cols=row.split('\t');
      const comentario=(cols[0]||'').trim();
      const n_uds   =parseNum(cols[1]||'0');
      const longitud=parseNum(cols[2]||'0');
      const anchura =parseNum(cols[3]||'0');
      const altura  =parseNum(cols[4]||'0');
      chain=chain.then(j=>{lastJ=j;return api({accion:'add_linea_medicion',
        codigo_hijo:cod,codigo_padre:pc,comentario,n_uds,longitud,anchura,altura})});
    });
    chain.then(j=>refresh(j||lastJ));
  };
  document.addEventListener('paste',_pasteHandler);
}

// ---- Actions ----
// Las funciones add* abren el capítulo destino y enfocan la ghost row,
// permitiendo al usuario rellenar código/descripción/ud/precio inline (no usan prompt).
function addPartida(codPadre){
  if(codPadre) openCaps.add(codPadre);
  renderTree();
  focusGhostRow(codPadre||'');
}
function addCapitulo(codPadre){
  if(codPadre) openCaps.add(codPadre);
  renderTree();
  focusGhostRow(codPadre||'');
}
async function eliminarConcepto(cod,codPadre){
  if(!confirm('¿Eliminar "'+cod+'"?'))return;
  refresh(await api({accion:'eliminar_concepto',codigo:cod,codigo_padre:codPadre}))
}
async function addLineaMed(codHijo,codPadre){
  refresh(await api({accion:'add_linea_medicion',codigo_hijo:codHijo,codigo_padre:codPadre,comentario:'',n_uds:1,longitud:0,anchura:0,altura:0}))
}
async function eliminarLineaMed(codHijo,codPadre,i){
  if(!confirm('¿Eliminar línea?'))return;
  refresh(await api({accion:'eliminar_linea_medicion',codigo_hijo:codHijo,codigo_padre:codPadre,indice:i}))
}

// ---- Informes ----
function descargarInforme(tipo){closeDropdowns();window.open('/api/informe?tipo='+tipo,'_blank')}
function exportarBC3(){window.location='/api/exportar'}
function toggleDropdown(btn){const m=btn.nextElementSibling;const o=m.classList.toggle('open');if(o)setTimeout(()=>document.addEventListener('click',function h(e){if(!m.contains(e.target)&&e.target!==btn){m.classList.remove('open');document.removeEventListener('click',h)}},0))}
function closeDropdowns(){document.querySelectorAll('.dropdown-menu.open').forEach(m=>m.classList.remove('open'))}
function esc(s){if(s==null)return'';const d=document.createElement('div');d.textContent=String(s);return d.innerHTML}
function showToast(msg,ms=2200){
  let t=document.getElementById('_toast');
  if(!t){t=document.createElement('div');t.id='_toast';t.style.opacity='0';document.body.appendChild(t)}
  t.textContent=msg;t.style.opacity='1';
  clearTimeout(t._tid);t._tid=setTimeout(()=>{t.style.opacity='0'},ms);
}

// Construye un TSV (tabulado) de una lista de nodos del árbol, listo para pegar
// en Excel: Código, Descripción, Ud, Cantidad, Precio, Importe (formato europeo).
function _construirTSV(nodos){
  const head=['Código','Descripción','Ud','Cantidad','Precio','Importe'].join('\t');
  const filas=nodos.map(n=>[
    n.codigo||'', n.resumen||'', n.unidad||'',
    n.medicion_fmt||'', n.precio_fmt||'', n.importe_fmt||''
  ].map(c=>String(c).replace(/\t/g,' ').replace(/\r?\n/g,' ')).join('\t'));
  return [head,...filas].join('\r\n');
}
// Escribe texto en el portapapeles del SISTEMA (para pegar fuera, p.ej. Excel).
// Funciona en contexto seguro (localhost cuenta como seguro).
function _escribirPortapapeles(texto){
  if(navigator.clipboard&&navigator.clipboard.writeText){
    navigator.clipboard.writeText(texto).catch(()=>_copiarFallback(texto));
  }else{
    _copiarFallback(texto);
  }
}
function _copiarSistemaTSV(nodos){
  if(!nodos||!nodos.length)return;
  _escribirPortapapeles(_construirTSV(nodos));
}
// Número → formato europeo (coma decimal) para pegar en Excel español.
function _numEU(x){
  if(x==null||x==='')return '';
  if(typeof x==='number')return String(x).replace('.',',');
  return String(x);
}
// Copia las filas de una tabla Tabulator (desglose o mediciones) como TSV.
// Si hay filas seleccionadas, copia esas; si no, copia todas.
function _copiarTablaTSV(tab, tipo){
  if(!tab)return;
  let rows=[];
  try{ rows=tab.getSelectedData(); }catch(e){}
  if(!rows||!rows.length){ try{ rows=tab.getData(); }catch(e){} }
  rows=(rows||[]).filter(r=>!r._isNew);
  if(!rows.length){ showToast('No hay filas que copiar'); return; }
  let head, filas;
  if(tipo==='descomp'){
    const TL={'1':'MO','2':'MQ','3':'MT','4':'AUX'};
    head=['Código','Tipo','Descripción','Ud','Rendimiento','Precio','Importe'];
    filas=rows.map(r=>[r.codigo,TL[r.tipo_fiebdc]||'',r.resumen,r.unidad,
      _numEU(r.rendimiento),_numEU(r.precio),r.importe_fmt||'']);
  }else{
    head=['Comentario','Uds','Largo','Ancho','Alto','Parcial'];
    filas=rows.map(r=>[r.comentario,_numEU(r.n_uds),_numEU(r.longitud),
      _numEU(r.anchura),_numEU(r.altura),r.subtotal_fmt||'']);
  }
  const limpia=c=>String(c==null?'':c).replace(/\t/g,' ').replace(/\r?\n/g,' ');
  const tsv=[head.join('\t'),...filas.map(f=>f.map(limpia).join('\t'))].join('\r\n');
  _escribirPortapapeles(tsv);
  showToast(`📋 Copiadas ${filas.length} fila${filas.length>1?'s':''} (pega en Excel)`);
}

// Elimina los recursos seleccionados del desglose de la partida actual.
async function eliminarSelDesglose(){
  if(!_tabDescomp||!curNode)return;
  const sel=_tabDescomp.getSelectedData().filter(r=>!r._isNew);
  if(!sel.length){showToast('Selecciona recursos (casilla izquierda)');return;}
  if(!confirm(`¿Quitar ${sel.length} recurso(s) de la descomposición?`))return;
  const cod=curNode.codigo;
  let lastJ=null;
  for(const r of sel){
    const j=await api({accion:'eliminar_recurso',codigo_partida:cod,codigo_recurso:r.codigo});
    if(j)lastJ=j;
  }
  if(lastJ)calcSave(null,lastJ);
}
// Elimina las líneas de medición seleccionadas (de mayor a menor índice para no desfasar).
async function eliminarSelMedic(){
  if(!_tabMedic||!curNode||!curParent)return;
  const sel=_tabMedic.getSelectedData().filter(r=>!r._isNew);
  if(!sel.length){showToast('Selecciona líneas (casilla izquierda)');return;}
  if(!confirm(`¿Eliminar ${sel.length} línea(s) de medición?`))return;
  const cod=curNode.codigo, pc=curParent;
  const idxs=sel.map(r=>r._idx).sort((a,b)=>b-a);   // descendente
  let lastJ=null;
  for(const idx of idxs){
    const j=await api({accion:'eliminar_linea_medicion',codigo_hijo:cod,codigo_padre:pc,indice:idx});
    if(j)lastJ=j;
  }
  if(lastJ)calcSave(null,lastJ);
}
// Fallback para navegadores/contextos sin Clipboard API.
function _copiarFallback(texto){
  try{
    const ta=document.createElement('textarea');
    ta.value=texto;ta.style.position='fixed';ta.style.opacity='0';
    document.body.appendChild(ta);ta.select();
    document.execCommand('copy');document.body.removeChild(ta);
  }catch(e){}
}

// Auto-load file if passed as argument
window.addEventListener('DOMContentLoaded', ()=>{
  // Restore saved theme
  const saved=localStorage.getItem('bc3theme');
  if(saved){document.documentElement.setAttribute('data-theme',saved);updateThemeBtn(saved)}
  else{document.documentElement.setAttribute('data-theme','dark');updateThemeBtn('dark')}

  fetch('/api/tiene_archivo').then(r=>r.json()).then(d=>{
    if(d.tiene){
      document.getElementById('loadingOverlay').style.display='';
      fetch('/api/cargar_local',{method:'POST'}).then(r=>r.json()).then(data=>{
        document.getElementById('loadingOverlay').style.display='none';
        if(data.error){alert(data.error);return}
        fileInfo=data.info;treeData=data.arbol;
        _discrepancias=data.discrepancias||[];
        _pemValidacion=data.pem||null;
        if(data.archivo)document.getElementById('fileName').textContent=data.archivo;
        renderApp();
        actualizarBannerDiscrepancias();
      })
    }
  })
})

async function undoAction(){
  const r=await fetch('/api/undo',{method:'POST'});
  const j=await r.json();
  if(j.error){showToast(j.error);return}
  refresh(j);
}
async function redoAction(){
  const r=await fetch('/api/redo',{method:'POST'});
  const j=await r.json();
  if(j.error){showToast(j.error);return}
  refresh(j);
}
// Rastrea en qué zona está trabajando el usuario, para que Ctrl+C copie lo
// correcto (árbol de partidas, desglose o mediciones).
document.addEventListener('mousedown',e=>{
  if(e.target.closest && e.target.closest('#tab-descomp')) _ctxCopia='descomp';
  else if(e.target.closest && e.target.closest('#tab-medic')) _ctxCopia='medic';
  else if(e.target.closest && e.target.closest('.tree-scroll')) _ctxCopia='arbol';
},true);

document.addEventListener('keydown',e=>{
  // Ignorar siempre si hay una celda en edición activa
  if(document.activeElement&&document.activeElement.contentEditable==='true')return;

  // Ctrl+Shift+Z o Ctrl+Y — rehacer
  if((e.ctrlKey||e.metaKey)&&((e.key==='z'&&e.shiftKey)||e.key==='y')){
    e.preventDefault();redoAction();
    return;
  }
  // Ctrl+Z — deshacer
  if((e.ctrlKey||e.metaKey)&&e.key==='z'&&!e.shiftKey){
    e.preventDefault();undoAction();
    return;
  }

  // Ctrl+C — copiar selección: portapapeles interno (pegar en la app) + TSV (pegar en Excel)
  if((e.ctrlKey||e.metaKey)&&e.key==='c'&&!e.shiftKey){
    if(window.getSelection&&window.getSelection().toString())return;
    // Según la zona con el foco, copia desglose o mediciones a Excel
    if(_ctxCopia==='descomp'&&_tabDescomp){e.preventDefault();_copiarTablaTSV(_tabDescomp,'descomp');return;}
    if(_ctxCopia==='medic'&&_tabMedic){e.preventDefault();_copiarTablaTSV(_tabMedic,'medic');return;}
    // Resto: copiar partidas del árbol
    if(!curNode)return;
    e.preventDefault();
    let nodos;
    if(selectedNodes.length>1){
      _clipboard=selectedNodes.map(s=>({codigo:s.codigo,padre:s.padre}));
      nodos=selectedNodes.map(s=>_findNodeInParent(treeData,s.codigo,s.padre)).filter(Boolean);
      showToast(`📋 Copiados ${selectedNodes.length} conceptos (Ctrl+V aquí o pega en Excel)`);
    }else{
      _clipboard={codigo:curNode.codigo,resumen:curNode.resumen};
      nodos=[curNode];
      document.querySelectorAll('.ttable tr.copied-row').forEach(r=>r.classList.remove('copied-row'));
      document.querySelectorAll('.ttable tbody tr.active').forEach(r=>r.classList.add('copied-row'));
      showToast(`📋 Copiado: ${curNode.codigo} (Ctrl+V aquí o pega en Excel)`);
    }
    _copiarSistemaTSV(nodos);   // también al portapapeles del sistema, para Excel
    return;
  }

  // Ctrl+V — pegar concepto(s) copiados justo debajo del concepto seleccionado
  if((e.ctrlKey||e.metaKey)&&e.key==='v'&&!e.shiftKey){
    if(!_clipboard||!curNode||!curParent)return;
    e.preventDefault();
    const antes_de=_siblingAfter(curParent,curNode.codigo);
    const items=Array.isArray(_clipboard)?_clipboard:[{codigo:_clipboard.codigo,padre:curParent}];
    const errs=[];let lastJ=null;
    // Pegar en serie, recoger errores parciales
    (async()=>{
      for(const it of items){
        const j=await fetch('/api/editar',{method:'POST',
          headers:{'Content-Type':'application/json'},
          body:JSON.stringify({accion:'copiar',codigo:it.codigo,padre_destino:curParent,antes_de})
        }).then(r=>r.json());
        if(j.error)errs.push(`${it.codigo}: ${j.error}`);
        else lastJ=j;
      }
      if(lastJ){refresh(lastJ);}
      if(errs.length)showToast('⚠ '+errs.join(' | '),4000);
      else showToast(`✓ Pegado${items.length>1?' ('+items.length+')':': '+items[0].codigo}`);
    })();
    return;
  }

  // Supr — eliminar lo seleccionado según la zona con el foco
  if(e.key==='Delete'&&!e.ctrlKey&&!e.metaKey&&!e.shiftKey){
    // En desglose o mediciones: borra las filas seleccionadas de esa tabla
    if(_ctxCopia==='descomp'&&_tabDescomp&&_tabDescomp.getSelectedData().filter(r=>!r._isNew).length){
      e.preventDefault();eliminarSelDesglose();return;
    }
    if(_ctxCopia==='medic'&&_tabMedic&&_tabMedic.getSelectedData().filter(r=>!r._isNew).length){
      e.preventDefault();eliminarSelMedic();return;
    }
    if(selectedNodes.length>1){
      e.preventDefault();
      const lista=selectedNodes.map(s=>s.codigo).join(', ');
      if(!confirm(`¿Eliminar ${selectedNodes.length} conceptos?\n${lista}`))return;
      (async()=>{
        let lastJ=null;
        for(const s of [...selectedNodes]){
          const j=await api({accion:'eliminar_concepto',codigo:s.codigo,codigo_padre:s.padre});
          if(j)lastJ=j;
        }
        selectedNodes=[];
        if(lastJ)refresh(lastJ);
      })();
      return;
    }
    if(!curNode||!curParent)return;
    e.preventDefault();
    eliminarConcepto(curNode.codigo,curParent);
  }
});

function toggleTheme(){
  const cur=document.documentElement.getAttribute('data-theme')||'dark';
  const next=cur==='dark'?'light':'dark';
  document.documentElement.setAttribute('data-theme',next);
  localStorage.setItem('bc3theme',next);
  updateThemeBtn(next)
}
function updateThemeBtn(theme){
  const btn=document.getElementById('themeBtn');
  if(btn)btn.textContent=theme==='dark'?'☀️':'🌙'
}
</script></body></html>"""

def main():
    import sys
    port = int(os.environ.get("PORT", 5000))

    # Si se pasa un archivo como argumento, cargarlo automáticamente al abrir
    if len(sys.argv) > 1:
        ruta = sys.argv[1]
        if os.path.isfile(ruta):
            _estado["ruta_arg"] = os.path.abspath(ruta)
            print(f"  Archivo: {ruta}")
        else:
            print(f"  Aviso: no se encuentra '{ruta}'")

    Timer(1.2, lambda: webbrowser.open(f"http://localhost:{port}")).start()
    print(f"\n  BC3Manager web -> http://localhost:{port}")
    print("  Los cambios se guardan automaticamente al archivo.")
    print("  Pulsa Ctrl+C para cerrar\n")
    app.run(host="127.0.0.1", port=port, debug=False)

if __name__ == "__main__":
    main()
